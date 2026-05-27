"""
==============================================================
  SISTEMA DE GLOSAS - APP WEB LOCAL
  Backend Flask - Orquesta DESCARGA y DESAPILAR
  Desarrollado por: DESARROLLO E INNOVACION SALUD NET
==============================================================
"""

import os
import re
import sys
import json
import queue
import shutil
import threading
import time
import zipfile
import io
from pathlib import Path
from datetime import datetime
from flask import Flask, request, jsonify, Response, send_file, stream_with_context
# flask_cors no requerido — CORS manejado manualmente
from werkzeug.utils import secure_filename

# ── Rutas base ──────────────────────────────────────────────
BASE_DIR     = Path(__file__).resolve().parent
UPLOAD_DIR   = BASE_DIR / "uploads"
OUTPUT_DIR   = BASE_DIR / "output"
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

app  = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024  # 200 MB

# CORS manual — no requiere flask-cors
@app.after_request
def add_cors_headers(response):
    response.headers["Access-Control-Allow-Origin"]  = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
    response.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
    return response

# ── Cola de logs por sesión ──────────────────────────────────
log_queues: dict[str, queue.Queue] = {}
job_status: dict[str, dict]        = {}


def get_queue(session_id: str) -> queue.Queue:
    if session_id not in log_queues:
        log_queues[session_id] = queue.Queue()
    return log_queues[session_id]


def push_log(session_id: str, msg: str, level: str = "info"):
    q = get_queue(session_id)
    q.put({"type": "log", "level": level, "msg": msg,
           "ts": datetime.now().strftime("%H:%M:%S")})


def push_done(session_id: str, payload: dict):
    q = get_queue(session_id)
    q.put({"type": "done", **payload})


# ══════════════════════════════════════════════════════════════
#  MÓDULO DESAPILAR  (wrapper sobre DESAPILAR_GLOSAS.py)
# ══════════════════════════════════════════════════════════════

def run_desapilar(pdf_paths: list[Path], work_dir: Path,
                  session_id: str) -> list[dict]:
    """
    Ejecuta la lógica de DESAPILAR_GLOSAS sobre una lista de PDFs.
    Retorna la lista de registros generados.
    """
    # Importar dinámicamente para que use work_dir como BASE_DIR
    import importlib.util, types

    push_log(session_id, f"Cargando módulo DESAPILAR_GLOSAS…")

    spec = importlib.util.spec_from_file_location(
        "desapilar_mod",
        Path(__file__).parent.parent / "DESAPILAR_GLOSAS.py"
        if not (BASE_DIR / "DESAPILAR_GLOSAS.py").exists()
        else BASE_DIR / "DESAPILAR_GLOSAS.py"
    )
    mod = importlib.util.module_from_spec(spec)

    # Parchear BASE_DIR e INPUT_DIR antes de ejecutar el módulo
    mod.__dict__["__file__"] = str(BASE_DIR / "DESAPILAR_GLOSAS.py")
    spec.loader.exec_module(mod)

    # Redirigir BASE_DIR del módulo al work_dir
    mod.BASE_DIR  = work_dir
    mod.INPUT_DIR = work_dir
    mod.MATRIX    = mod.load_matrix.__func__(work_dir) if hasattr(mod.load_matrix, "__func__") else _reload_matrix(mod, work_dir)

    rows = []
    total = len(pdf_paths)

    for idx, pdf_path in enumerate(pdf_paths, 1):
        push_log(session_id, f"[{idx}/{total}] Procesando: {pdf_path.name}")
        try:
            mod.process_pdf(pdf_path, rows)
        except Exception as e:
            push_log(session_id, f"  ✗ Error en {pdf_path.name}: {e}", "error")

    return rows


def _reload_matrix(mod, work_dir: Path):
    """Recarga la MATRIZ_IPS buscándola en work_dir o en BASE_DIR del módulo."""
    import pandas as pd
    orig_base = mod.BASE_DIR
    mod.BASE_DIR = work_dir
    try:
        result = mod.load_matrix()
    except Exception:
        mod.BASE_DIR = orig_base
        result = mod.load_matrix()
    mod.BASE_DIR = work_dir
    return result


def desapilar_standalone(pdf_paths: list[Path], session_id: str) -> dict:
    """
    Desapila PDFs subidos manualmente.
    Crea un work_dir temporal, copia los PDFs, corre desapilar y empaqueta.
    """
    import importlib.util

    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    work_dir = OUTPUT_DIR / f"desapilar_{ts}"
    work_dir.mkdir(parents=True, exist_ok=True)

    # Copiar PDFs al work_dir
    for p in pdf_paths:
        shutil.copy2(p, work_dir / p.name)

    push_log(session_id, f"Work dir: {work_dir.name}")
    push_log(session_id, f"PDFs a procesar: {len(pdf_paths)}")

    # Verificar que DESAPILAR_GLOSAS.py esté accesible
    desapilar_py = _find_desapilar_py()
    if not desapilar_py:
        push_log(session_id, "DESAPILAR_GLOSAS.py no encontrado. Asegúrese de que esté en la misma carpeta.", "error")
        return {"ok": False, "error": "DESAPILAR_GLOSAS.py no encontrado"}

    rows = _run_desapilar_direct(desapilar_py, work_dir, pdf_paths, session_id)

    # Empaquetar resultados
    zip_path = _pack_results(work_dir, session_id)
    report   = _get_report(work_dir)

    return {
        "ok":       True,
        "rows":     rows,
        "zip":      str(zip_path.relative_to(BASE_DIR)) if zip_path else None,
        "report":   str(report.relative_to(BASE_DIR)) if report else None,
        "work_dir": str(work_dir),
    }


def _find_desapilar_py() -> Path | None:
    candidates = [
        BASE_DIR / "DESAPILAR_GLOSAS.py",
        BASE_DIR.parent / "DESAPILAR_GLOSAS.py",
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def _find_descarga_py() -> Path | None:
    candidates = [
        BASE_DIR / "DESCARGA_GLOSAS.py",
        BASE_DIR.parent / "DESCARGA_GLOSAS.py",
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def _find_unificar_py() -> Path | None:
    candidates = [
        BASE_DIR / "unificar_pdfs.py",
        BASE_DIR.parent / "unificar_pdfs.py",
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def _find_renombrar_py() -> Path | None:
    candidates = [
        BASE_DIR / "renombrar_auditoria.py",
        BASE_DIR.parent / "renombrar_auditoria.py",
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def _find_bolivar_py() -> Path | None:
    candidates = [
        BASE_DIR / "DESCARGA_BOLIVAR.py",
        BASE_DIR.parent / "DESCARGA_BOLIVAR.py",
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def _find_portal_py() -> Path | None:
    candidates = [
        BASE_DIR / "DESCARGA_PORTAL.py",
        BASE_DIR.parent / "DESCARGA_PORTAL.py",
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def _run_desapilar_direct(desapilar_py: Path, work_dir: Path,
                           pdf_paths: list[Path], session_id: str) -> list[dict]:
    """
    Importa DESAPILAR_GLOSAS con BASE_DIR = work_dir y procesa los PDFs.
    """
    import importlib.util

    spec = importlib.util.spec_from_file_location("desapilar_dyn", str(desapilar_py))
    mod  = importlib.util.module_from_spec(spec)

    # Inyectar BASE_DIR antes de exec para que load_matrix() lo use
    sys.modules["desapilar_dyn"] = mod
    spec.loader.exec_module(mod)

    # Redirigir rutas
    mod.BASE_DIR  = work_dir
    mod.INPUT_DIR = work_dir

    # Recargar matrix apuntando al work_dir (busca MATRIZ_IPS.xlsx allí o en app_dir)
    push_log(session_id, "Cargando MATRIZ_IPS…")
    for search_dir in [work_dir, BASE_DIR, BASE_DIR.parent]:
        matrix_path = search_dir / "MATRIZ_IPS.xlsx"
        if matrix_path.exists():
            # Copiar al work_dir si no está allí
            if search_dir != work_dir:
                shutil.copy2(matrix_path, work_dir / "MATRIZ_IPS.xlsx")
            break

    mod.MATRIX = mod.load_matrix()
    if not mod.MATRIX:
        push_log(session_id, "⚠ Matriz vacía — verifique MATRIZ_IPS.xlsx", "warn")

    rows  = []
    total = len(pdf_paths)
    for idx, p in enumerate(pdf_paths, 1):
        wp = work_dir / p.name
        push_log(session_id, f"[{idx}/{total}] {p.name}")
        try:
            mod.process_pdf(wp, rows)
            push_log(session_id, f"  ✓ {p.name} → {len(rows)} registros acumulados")
        except Exception as e:
            push_log(session_id, f"  ✗ {p.name}: {e}", "error")

    # Guardar reporte Excel explicitamente
    if rows:
        try:
            import pandas as pd
            report_path = work_dir / "REPORTE_GLOSAS.xlsx"
            pd.DataFrame(rows, columns=[
                "factura procesada", "aseguradora",
                "documento original", "ruta destino final",
            ]).to_excel(report_path, index=False)
            push_log(session_id, f"Reporte Excel guardado: {len(rows)} registros")
        except Exception as e:
            push_log(session_id, f"Error al guardar reporte: {e}", "warn")

    return rows


def _pack_results(work_dir: Path, session_id: str) -> Path | None:
    """Comprime todos los PDFs generados (excluye PROCESADOS y PROCESAR MANUAL)."""
    zip_path = work_dir / "resultado_desapilar.zip"
    count    = 0
    push_log(session_id, "Empaquetando resultados…")
    try:
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in work_dir.rglob("*.pdf"):
                # Excluir los PDFs originales en PROCESADOS
                rel = f.relative_to(work_dir)
                parts = rel.parts
                if parts[0] in ("PROCESADOS", "uploads"):
                    continue
                zf.write(f, rel)
                count += 1
        push_log(session_id, f"ZIP creado con {count} archivo(s)")
        return zip_path
    except Exception as e:
        push_log(session_id, f"Error al crear ZIP: {e}", "error")
        return None


def _get_report(work_dir: Path) -> Path | None:
    r = work_dir / "REPORTE_GLOSAS.xlsx"
    return r if r.exists() else None


# ══════════════════════════════════════════════════════════════
#  MÓDULO DESCARGA  (wrapper sobre DESCARGA_GLOSAS.py)
# ══════════════════════════════════════════════════════════════

def run_descarga_y_desapilar(config: dict, session_id: str) -> dict:
    """
    1. Ejecuta DESCARGA_GLOSAS para bajar PDFs del correo.
    2. Toma los PDFs descargados y los pasa por DESAPILAR_GLOSAS.
    3. Retorna registros combinados.
    """
    import importlib.util

    descarga_py = _find_descarga_py()
    if not descarga_py:
        return {"ok": False, "error": "DESCARGA_GLOSAS.py no encontrado"}

    desapilar_py = _find_desapilar_py()
    if not desapilar_py:
        return {"ok": False, "error": "DESAPILAR_GLOSAS.py no encontrado"}

    ts           = datetime.now().strftime("%Y%m%d_%H%M%S")
    work_dir     = OUTPUT_DIR / f"descarga_{ts}"
    download_dir = work_dir / "DESCARGA DESDE CORREOS"
    work_dir.mkdir(parents=True, exist_ok=True)
    download_dir.mkdir(parents=True, exist_ok=True)

    # ── Paso 1: Descarga ────────────────────────────────────
    push_log(session_id, "═══ FASE 1: DESCARGA DESDE CORREO ═══")

    spec = importlib.util.spec_from_file_location("descarga_dyn", str(descarga_py))
    mod_d = importlib.util.module_from_spec(spec)
    sys.modules["descarga_dyn"] = mod_d
    spec.loader.exec_module(mod_d)

    # Redirigir rutas de descarga
    # EXCEL_CONFIG apunta a donde realmente está el archivo (carpeta padre o base)
    excel_config_path = None
    for search in [BASE_DIR, BASE_DIR.parent, BASE_DIR.parent.parent]:
        candidate = search / "NOMBRES Y NIT EQUIVALENTES.xlsx"
        if candidate.exists():
            excel_config_path = str(candidate)
            break
    excel_asuntos_path = None
    for search in [BASE_DIR, BASE_DIR.parent, BASE_DIR.parent.parent]:
        candidate = search / "asuntos_correos_glosas.xlsx"
        if candidate.exists():
            excel_asuntos_path = str(candidate)
            break

    mod_d.DIRECTORIO_BASE = str(work_dir)
    mod_d.RUTA_BASE       = str(download_dir)
    mod_d.CARPETA_NO_ID   = str(download_dir / "_NO_IDENTIFICADOS")
    mod_d.REPORTE_EXCEL   = str(work_dir / "REPORTE_DESCARGA.xlsx")
    mod_d.EXCEL_CONFIG    = excel_config_path or config.get("excel_config", "")
    mod_d.EXCEL_ASUNTOS   = excel_asuntos_path or config.get("excel_asuntos", "")

    # Cuentas vienen siempre del formulario
    cuentas_override = config.get("cuentas") or []
    if not cuentas_override:
        push_log(session_id, "No se ingresaron cuentas. Proceso cancelado.", "error")
        return {"ok": False, "error": "Debe ingresar al menos una cuenta Gmail en el formulario."}

    registros_descarga = []
    push_log(session_id, "Cargando configuración de IPS y aseguradoras…")
    try:
        _, ips_dict, aseg_dict, patrones, aseg_passwords = mod_d.cargar_configuracion()
        cuentas = cuentas_override
        push_log(session_id, f"Usando {len(cuentas)} cuenta(s) del formulario")

        sesion_id_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        for idx, cuenta in enumerate(cuentas, 1):
            push_log(session_id, f"[{idx}/{len(cuentas)}] Procesando: {cuenta['correo']}")
            try:
                regs, ok, dup, no_id = mod_d.procesar_cuenta(
                    cuenta, ips_dict, aseg_dict, patrones, sesion_id_str
                )
                registros_descarga.extend(regs)
                push_log(session_id,
                         f"  ✓ OK:{ok} Duplicados:{dup} NoID:{no_id}")
            except Exception as e:
                push_log(session_id, f"  ✗ {cuenta['correo']}: {e}", "error")

    except Exception as e:
        push_log(session_id, f"Error en descarga: {e}", "error")
        return {"ok": False, "error": str(e)}

    # ── Paso 2: Recolectar PDFs descargados ─────────────────
    push_log(session_id, "═══ FASE 2: RECOLECTANDO PDFs DESCARGADOS ═══")
    pdfs_descargados = list(download_dir.rglob("*.pdf"))
    # Excluir NO_IDENTIFICADOS si se desea — incluirlos igual para desapilar
    push_log(session_id, f"PDFs descargados: {len(pdfs_descargados)}")

    if not pdfs_descargados:
        push_log(session_id, "No se descargaron PDFs. Proceso terminado.", "warn")
        return {
            "ok":                True,
            "registros_descarga": registros_descarga,
            "registros_desapilar": [],
            "zip":               None,
            "work_dir":          str(work_dir),
        }

    # ── Paso 3: Desapilar los PDFs descargados ──────────────
    push_log(session_id, "═══ FASE 3: DESAPILAR PDFs ═══")

    # Copiar PDFs al input_dir con nombres únicos para evitar colisiones
    desapilar_input_dir = work_dir / "_desapilar_input"
    desapilar_input_dir.mkdir(exist_ok=True)

    # Mapeo original -> destino (nombre único)
    pdf_map = {}
    name_count = {}
    for p in pdfs_descargados:
        stem, suffix = p.stem, p.suffix
        count = name_count.get(p.name, 0)
        name_count[p.name] = count + 1
        dest_name = p.name if count == 0 else f"{stem}_{count}{suffix}"
        dest = desapilar_input_dir / dest_name
        shutil.copy2(p, dest)
        pdf_map[p] = dest

    # Copiar MATRIZ_IPS si existe (prefijos de facturas) — buscar en 3 niveles
    for search_dir in [BASE_DIR, BASE_DIR.parent, BASE_DIR.parent.parent]:
        mx = search_dir / "MATRIZ_IPS.xlsx"
        if mx.exists():
            shutil.copy2(mx, desapilar_input_dir / "MATRIZ_IPS.xlsx")
            break

    # Copiar NOMBRES_Y_NIT_EQUIVALENTES si existe — buscar en 3 niveles
    for search_dir in [BASE_DIR, BASE_DIR.parent, BASE_DIR.parent.parent]:
        for nombre_archivo in ["NOMBRES Y NIT EQUIVALENTES.xlsx",
                               "NOMBRES_Y_NIT_EQUIVALENTES.xlsx"]:
            nx = search_dir / nombre_archivo
            if nx.exists():
                shutil.copy2(nx, desapilar_input_dir / "NOMBRES Y NIT EQUIVALENTES.xlsx")
                break
        else:
            continue
        break

    spec2 = importlib.util.spec_from_file_location("desapilar_dyn2", str(desapilar_py))
    mod_ds = importlib.util.module_from_spec(spec2)
    sys.modules["desapilar_dyn2"] = mod_ds
    spec2.loader.exec_module(mod_ds)

    mod_ds.BASE_DIR  = desapilar_input_dir
    mod_ds.INPUT_DIR = desapilar_input_dir
    mod_ds.MATRIX    = mod_ds.load_matrix()

    rows_desapilar = []
    total = len(pdf_map)
    for idx, (orig, wp) in enumerate(pdf_map.items(), 1):
        push_log(session_id, f"  Desapilando [{idx}/{total}]: {orig.name}")
        try:
            mod_ds.process_pdf(wp, rows_desapilar)
        except Exception as e:
            push_log(session_id, f"  ✗ {orig.name}: {e}", "error")

    # ── Reorganizar: mover desapilados a CIUDAD/IPS/AÑO/MES/DIA/ASEGURADORA ──────
    # El desapilador genera: _desapilar_input/ASEGURADORA/factura.pdf
    #                    o:  _desapilar_input/PROCESAR MANUAL/factura.pdf
    # El original estaba en: download_dir/CIUDAD/IPS/AÑO/MES/DIA/ASEG_ORIG/orig.pdf
    # Queremos:              download_dir/CIUDAD/IPS/AÑO/MES/DIA/ASEGURADORA/factura.pdf
    push_log(session_id, "Reorganizando archivos desapilados en estructura de carpetas...")

    # Construir indice: nombre_wp -> carpeta hasta DIA
    # Estructura download: CIUDAD / NIT-IPS / AÑO / MES / DIA / ASEGURADORA / pdf
    wp_to_base = {}
    for orig_pdf, wp in pdf_map.items():
        # orig_pdf.parent = .../ASEGURADORA, .parent = .../DIA -> subir dos niveles para DIA
        dia_folder = orig_pdf.parent.parent.parent
        wp_to_base[wp.name] = dia_folder

    # Recopilar TODOS los PDFs generados por el desapilador en desapilar_input_dir
    # incluyendo subcarpetas ASEGURADORA y PROCESAR MANUAL
    todos_generados = list(desapilar_input_dir.rglob("*.pdf"))
    # Excluir los originales que copiamos al input
    nombres_wp = {wp.name for wp in pdf_map.values()}
    todos_generados = [f for f in todos_generados
                       if f.name not in nombres_wp
                       and "PROCESADOS" not in str(f)]

    nombres_originales = nombres_wp

    for pdf_generado in todos_generados:
        # La aseguradora es el nombre de la carpeta padre dentro de desapilar_input_dir
        try:
            rel = pdf_generado.relative_to(desapilar_input_dir)
            aseg = rel.parts[0] if len(rel.parts) > 1 else "_ASEGURADORA_NO_IDENTIFICADA"
        except Exception:
            aseg = "_ASEGURADORA_NO_IDENTIFICADA"

        # Buscar qué original generó este PDF usando rows_desapilar
        base_folder = None
        for row in rows_desapilar:
            ruta_dest = row.get("ruta destino final", "")
            if ruta_dest and Path(ruta_dest).name == pdf_generado.name:
                doc_orig = row.get("documento original", "")
                for wp_name, dia_f in wp_to_base.items():
                    if wp_name in doc_orig or doc_orig in wp_name:
                        base_folder = dia_f
                        break
                break

        if base_folder is None:
            # Fallback: usar primera entrada disponible
            base_folder = next(iter(wp_to_base.values()), download_dir)

        dest_dir = base_folder / aseg
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest_file = dest_dir / pdf_generado.name
        try:
            if not dest_file.exists():
                shutil.copy2(str(pdf_generado), str(dest_file))
        except Exception as e:
            push_log(session_id, f"  ✗ No se pudo copiar {pdf_generado.name}: {e}", "warn")

    # Eliminar PDFs originales apilados de la carpeta de descarga
    for orig_pdf in pdf_map.keys():
        try:
            if orig_pdf.exists():
                orig_pdf.unlink()
        except Exception:
            pass

    # Limpiar carpetas de aseguradora originales que quedaron vacías
    for orig_pdf in pdf_map.keys():
        try:
            orig_folder = orig_pdf.parent
            if orig_folder.exists() and not any(orig_folder.iterdir()):
                orig_folder.rmdir()
        except Exception:
            pass

    # Guardar reporte combinado
    import pandas as pd
    if rows_desapilar:
        df = pd.DataFrame(rows_desapilar, columns=[
            "factura procesada", "aseguradora",
            "documento original", "ruta destino final",
        ])
        df.to_excel(work_dir / "REPORTE_GLOSAS.xlsx", index=False)

    # Empaquetar desde download_dir — ya tiene la estructura final correcta
    push_log(session_id, "Empaquetando resultado final...")
    zip_path = work_dir / "resultado_final.zip"
    import zipfile as zf_mod
    count_zip = 0
    try:
        with zf_mod.ZipFile(zip_path, "w", zf_mod.ZIP_DEFLATED) as zf:
            # Empaquetar desde download_dir (estructura CIUDAD/IPS/AÑO/MES/DIA/ASEG)
            for f in download_dir.rglob("*.pdf"):
                if f.name not in nombres_originales:
                    rel = f.relative_to(download_dir)
                    zf.write(f, rel)
                    count_zip += 1
            # Incluir también lo que quedó en PROCESAR MANUAL dentro del desapilar_input
            manual_dir = desapilar_input_dir / "PROCESAR MANUAL"
            if manual_dir.exists():
                for f in manual_dir.rglob("*.pdf"):
                    zf.write(f, Path("PROCESAR MANUAL") / f.name)
                    count_zip += 1
            # Incluir ZIPs de Claimonline guardados en ruta de IPS (ZIP_COLPATRIA)
            for f in download_dir.rglob("*.zip"):
                if "ZIP_COLPATRIA" in str(f):
                    rel = f.relative_to(download_dir)
                    zf.write(f, rel)
                    count_zip += 1
        push_log(session_id, f"ZIP creado con {count_zip} archivo(s)")
        if count_zip == 0:
            zip_path = None
    except Exception as e:
        push_log(session_id, f"Error al crear ZIP: {e}", "warn")
        zip_path = None

    push_log(session_id, "✓ Proceso completo", "success")
    return {
        "ok":                  True,
        "registros_descarga":  registros_descarga,
        "registros_desapilar": rows_desapilar,
        "zip":    str(zip_path.relative_to(BASE_DIR)) if zip_path and zip_path.exists() else None,
        "report": str((work_dir / "REPORTE_GLOSAS.xlsx").relative_to(BASE_DIR))
                  if (work_dir / "REPORTE_GLOSAS.xlsx").exists() else None,
        "work_dir": str(work_dir),
    }


# ══════════════════════════════════════════════════════════════
#  MÓDULO UNIFICAR  (wrapper sobre unificar_pdfs.py)
# ══════════════════════════════════════════════════════════════

def run_unificar(pdf_paths: list[Path], session_id: str) -> dict:
    """
    Recibe una lista de PDFs subidos, los copia a un work_dir y ejecuta
    la lógica de unificar_pdfs.py:
      - PDFs con la misma base numérica se fusionan en orden
      - PDFs únicos se renombran limpiamente
    Retorna ZIP con los PDFs resultantes y un resumen de registros.
    """
    import importlib.util

    unificar_py = _find_unificar_py()
    if not unificar_py:
        return {"ok": False, "error": "unificar_pdfs.py no encontrado"}

    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    work_dir = OUTPUT_DIR / f"unificar_{ts}"
    carpeta  = work_dir / "UNIFICAR - CORTE DIGITAL"
    carpeta.mkdir(parents=True, exist_ok=True)

    # Copiar PDFs al subdirectorio que espera el script
    for p in pdf_paths:
        shutil.copy2(p, carpeta / p.name)

    push_log(session_id, f"Work dir: {work_dir.name}")
    push_log(session_id, f"PDFs recibidos: {len(pdf_paths)}")

    # Cargar módulo dinámicamente y redirigir script_dir
    spec = importlib.util.spec_from_file_location("unificar_dyn", str(unificar_py))
    mod  = importlib.util.module_from_spec(spec)
    import sys as _sys
    _sys.modules["unificar_dyn"] = mod
    spec.loader.exec_module(mod)

    # Monkey-patch script_dir para que apunte al work_dir
    mod.script_dir = lambda: work_dir

    # Ejecutar main() capturando stdout como logs SSE
    import io as _io, contextlib
    buf = _io.StringIO()
    try:
        with contextlib.redirect_stdout(buf):
            mod.main()
    except SystemExit:
        pass
    except Exception as e:
        push_log(session_id, f"Error en unificar: {e}", "error")

    # Reenviar líneas del stdout como logs
    for line in buf.getvalue().splitlines():
        line = line.strip()
        if not line or line.startswith("="):
            continue
        level = "error" if "✘" in line or "ERROR" in line else \
                "warn"  if "OMITIDO" in line else \
                "success" if "✔" in line or "RENOMBRADO" in line else "info"
        push_log(session_id, line, level)

    # Recopilar resultados
    pdfs_resultado = list(carpeta.glob("*.pdf"))
    push_log(session_id, f"PDFs resultado: {len(pdfs_resultado)}")

    # Construir registros para la tabla
    rows = []
    for p in sorted(pdfs_resultado):
        rows.append({
            "factura procesada":   p.stem,
            "aseguradora":         "—",
            "documento original":  p.name,
            "ruta destino final":  str(p),
        })

    # Empaquetar en ZIP
    zip_path = work_dir / "resultado_unificar.zip"
    try:
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for p in pdfs_resultado:
                zf.write(p, p.name)
        push_log(session_id, f"ZIP creado con {len(pdfs_resultado)} archivo(s)", "success")
    except Exception as e:
        push_log(session_id, f"Error al crear ZIP: {e}", "error")
        zip_path = None

    return {
        "ok":   True,
        "rows": rows,
        "zip":  str(zip_path.relative_to(BASE_DIR)) if zip_path and zip_path.exists() else None,
    }


# ══════════════════════════════════════════════════════════════
#  MÓDULO RENOMBRAR  (wrapper sobre renombrar_auditoria.py)
# ══════════════════════════════════════════════════════════════

def run_renombrar(pdf_paths: list[Path], session_id: str) -> dict:
    """
    Recibe una lista de PDFs, los copia a un work_dir y ejecuta
    renombrar_auditoria.py:
      - Detecta aseguradora y número de factura en el texto del PDF
      - Renombra según el formato de cada aseguradora
      - Los no identificados van a la subcarpeta revision_manual/
    Retorna ZIP con procesados + revision_manual, reporte Excel y registros.
    """
    import importlib.util

    renombrar_py = _find_renombrar_py()
    if not renombrar_py:
        return {"ok": False, "error": "renombrar_auditoria.py no encontrado"}

    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    work_dir = OUTPUT_DIR / f"renombrar_{ts}"
    work_dir.mkdir(parents=True, exist_ok=True)

    # Copiar PDFs al work_dir (es el cwd que espera el script)
    for p in pdf_paths:
        shutil.copy2(p, work_dir / p.name)

    push_log(session_id, f"Work dir: {work_dir.name}")
    push_log(session_id, f"PDFs recibidos: {len(pdf_paths)}")

    # Cargar módulo dinámicamente
    spec = importlib.util.spec_from_file_location("renombrar_dyn", str(renombrar_py))
    mod  = importlib.util.module_from_spec(spec)
    import sys as _sys
    _sys.modules["renombrar_dyn"] = mod
    spec.loader.exec_module(mod)

    # Ejecutar procesamiento sobre cada PDF
    revision_dir = work_dir / "revision_manual"
    revision_dir.mkdir(exist_ok=True)

    rows   = []
    ok_n   = rev_n = err_n = 0

    total = len(pdf_paths)
    for idx, src in enumerate(pdf_paths, 1):
        dest_pdf = work_dir / src.name
        push_log(session_id, f"[{idx}/{total}] {src.name}")
        try:
            estado, original, nuevo = mod.procesar_pdf(dest_pdf, work_dir, revision_dir)
            rows.append({
                "estado":            estado,
                "archivo_original":  original,
                "archivo_nuevo":     nuevo,
            })
            level = "success" if estado == "OK" else "warn" if estado == "REVISION" else "error"
            push_log(session_id,
                     f"  {'✓' if estado=='OK' else '⚠' if estado=='REVISION' else '✗'} "
                     f"{original} → {nuevo or '(revisión manual)'}",
                     level)
            if estado == "OK":       ok_n  += 1
            elif estado == "REVISION": rev_n += 1
            else:                    err_n += 1
        except Exception as e:
            push_log(session_id, f"  ✗ {src.name}: {e}", "error")
            rows.append({"estado": "ERROR", "archivo_original": src.name, "archivo_nuevo": str(e)})
            err_n += 1

    push_log(session_id, f"OK: {ok_n} | Revisión: {rev_n} | Error: {err_n}")

    # Guardar reporte Excel
    report_path = work_dir / "REPORTE_RENOMBRAR.xlsx"
    try:
        import pandas as pd
        pd.DataFrame(rows, columns=["estado", "archivo_original", "archivo_nuevo"]
                     ).to_excel(report_path, index=False)
        push_log(session_id, f"Reporte guardado: {len(rows)} registro(s)")
    except Exception as e:
        push_log(session_id, f"Error al guardar reporte: {e}", "warn")
        report_path = None

    # Empaquetar en ZIP (renombrados + revision_manual/)
    zip_path = work_dir / "resultado_renombrar.zip"
    count_zip = 0
    try:
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for p in work_dir.glob("*.pdf"):
                zf.write(p, p.name)
                count_zip += 1
            for p in revision_dir.glob("*.pdf"):
                zf.write(p, f"revision_manual/{p.name}")
                count_zip += 1
        push_log(session_id, f"ZIP creado con {count_zip} archivo(s)", "success")
    except Exception as e:
        push_log(session_id, f"Error al crear ZIP: {e}", "error")
        zip_path = None

    return {
        "ok":     True,
        "rows":   rows,
        "zip":    str(zip_path.relative_to(BASE_DIR)) if zip_path and zip_path.exists() else None,
        "report": str(report_path.relative_to(BASE_DIR)) if report_path and report_path.exists() else None,
    }


# ══════════════════════════════════════════════════════════════
#  MÓDULO BOLÍVAR SOAT  (Activa IT — descarga con Playwright)
# ══════════════════════════════════════════════════════════════

try:
    import openpyxl as _openpyxl
    _EXCEL_OK = True
except ImportError:
    _EXCEL_OK = False

# ── Mapa NIT → nombre IPS ────────────────────────────────────
_MAPA_IPS_BOLIVAR = {
    "900267064": "INVERSIONES_AZALUD_CLINICA_BAHIA",
    "900827065": "CENTRO_DE_DIAGNOSTICO_E_IMAGENES_BAHIA",
    "900657731": "CENTRO_MEDICO_Y_DE_REHABILITACION_BAHIA",
    "900826509": "RED_DE_URGENCIAS_DEL_MAGDALENA",
    "900513306": "FUNDACION_MARIA_REINA",
    "900600550": "INVERSIONES_MEDICAS_BARU",
    "900954800": "CENTRO_MEDICO_Y_DE_REHABILITACION_BARU",
    "900631361": "INVERSIONES_MEDICAS_VALLESALUD",
    "900257333": "ODONTOTRANS",
    "901081281": "URGETRAUMA",
    "900792417": "RED_DE_URGENCIAS_DE_LA_COSTA_PACIFICA",
    "901959993": "CLINICA_CORDIALIDAD",
    "900002780": "FUNDACION_CAMPBELL",
    "901523868": "MOVID_IPS_SAS",
    "901057487": "TECNOLOGIA_DIAGNOSTICA_DEL_VALLE",
    "900558595": "FUNDACION_MEDICA_CAMPBELL",
    "901149757": "UNIDAD_MEDICA_DE_TRAUMA_VALLE_SALUD",
    "900900754": "CLINICA_VALLE_SALUD_SAN_FERNANDO",
    "900469882": "CENTRO_MEDICO_SERVISALUD_INTEGRAL_IPS_SAS",
    "802024329": "RED_DE_URGENCIA_DE_LA_COSTA_LTDA",
    "900847382": "CENTRO_MEDICO_Y_DE_REHABILITACION_VALLE_SALUD",
}

# ── Estado global del job Bolívar ────────────────────────────
_bolivar_state: dict = {
    "running": False, "stopping": False, "finished": False,
    "error": None, "logs": [],
    "stats": {"total": 0, "descargadas": 0, "errores": 0},
    "errores_detalle": [], "descargas_exitosas": [],
    "facturas_permitidas": [],
}
_bolivar_lock   = threading.Lock()
_bolivar_browser  = None
_bolivar_context  = None
_bolivar_dl_dir   = None
_bolivar_periodo  = None
_bolivar_ips_name = None

_MESES_BOL = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]

BOLIVAR_DOWNLOAD_DIR = BASE_DIR / "downloads_bolivar"
BOLIVAR_DOWNLOAD_DIR.mkdir(exist_ok=True)


def _bol_log(msg: str, level: str = "info"):
    ts = datetime.now().strftime("%H:%M:%S")
    with _bolivar_lock:
        _bolivar_state["logs"].append({"ts": ts, "msg": msg, "level": level})


def _bol_validar_periodo(p: str) -> bool:
    if not p or len(p) < 5:
        return False
    return p[:3] in _MESES_BOL and bool(re.match(r'^\d{2}$', p[3:]))


def _bol_generar_rango(inicio: str, fin: str) -> list[str]:
    if not _bol_validar_periodo(inicio) or not _bol_validar_periodo(fin):
        return []
    mi, ai = _MESES_BOL.index(inicio[:3]), int(inicio[3:])
    mf, af = _MESES_BOL.index(fin[:3]),   int(fin[3:])
    if af * 100 + mf < ai * 100 + mi:
        return []
    result, mes, anio = [], mi, ai
    while True:
        result.append(_MESES_BOL[mes] + str(anio).zfill(2))
        if anio == af and mes == mf:
            break
        mes += 1
        if mes > 11:
            mes = 0
            anio += 1
    return result


def _bol_parse_periodo(raw: str) -> list[str]:
    raw = raw.strip()
    if not raw:
        return []
    # Formato YYYY-MM
    m = re.match(r'^(\d{4})-(\d{1,2})$', raw)
    if m:
        anio, mes = m.group(1)[2:], int(m.group(2)) - 1
        if 0 <= mes < 12:
            return [_MESES_BOL[mes] + anio]
    # Rango con guión (Abr26-Jun26)
    if '-' in raw:
        parts = [p.strip() for p in raw.split('-')]
        if len(parts) == 2:
            return _bol_generar_rango(parts[0], parts[1])
    # Formato directo MmmYY
    if _bol_validar_periodo(raw):
        return [raw]
    # Capitalizar (ene26 → Ene26)
    m2 = re.match(r'^([a-zA-Z]{3})(\d{2,4})$', raw)
    if m2:
        p = m2.group(1).capitalize() + m2.group(2)[-2:]
        if _bol_validar_periodo(p):
            return [p]
    return []


def _bol_cargar_progreso(ips_dir: Path) -> set:
    f = ips_dir / "progreso.json"
    if f.exists():
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            c = data.get("completadas", [])
            return set(c) if isinstance(c, list) else set()
        except Exception:
            pass
    return set()


def _bol_guardar_progreso(ips_dir: Path, completadas: set):
    try:
        (ips_dir / "progreso.json").write_text(
            json.dumps({"completadas": list(completadas),
                        "actualizado": datetime.now().isoformat()},
                       indent=2),
            encoding="utf-8",
        )
    except Exception as e:
        _bol_log(f"⚠ Error al guardar progreso: {e}", "warn")


def _bol_generar_excel(dl_dir: Path, periodo: str, ips_nombre: str,
                       exitosas: list, errores: list) -> Path | None:
    if not _EXCEL_OK:
        return None
    excel_path = dl_dir / ips_nombre / f"reporte_{periodo}.xlsx"
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.title = "Descargadas"
    ws.append(["N° Factura", "Estado", "IPS", "Archivo", "Fecha/Hora"])
    for ex in exitosas:
        ws.append([ex.get("factura"), ex.get("estado"), ips_nombre,
                   ex.get("archivo"), ex.get("timestamp")])
    ws2 = wb.create_sheet("Errores")
    ws2.append(["N° Factura", "Estado", "IPS", "Error", "Captura", "Fecha/Hora"])
    for err in errores:
        ws2.append([err.get("factura"), err.get("estado"), ips_nombre,
                    err.get("error"), err.get("captura"), err.get("timestamp")])
    wb.save(excel_path)
    return excel_path


def _bol_crear_zip(dl_dir: Path, periodo: str, ips_nombre: str) -> str | None:
    try:
        zip_path = dl_dir / f"facturas_{periodo}.zip"
        ips_dir  = dl_dir / ips_nombre
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            if ips_dir.exists():
                for pdf in ips_dir.rglob("*.pdf"):
                    zf.write(pdf, str(pdf.relative_to(dl_dir)))
                for excel in ips_dir.glob("reporte_*.xlsx"):
                    if "_PARCIAL_" not in excel.name:
                        zf.write(excel, str(excel.relative_to(dl_dir)))
                errores_dir = ips_dir / "Errores"
                if errores_dir.exists():
                    for f in errores_dir.rglob("*"):
                        zf.write(f, str(f.relative_to(dl_dir)))
        _bol_log(f"📦 ZIP creado: {zip_path.name}")
        return str(zip_path)
    except Exception as e:
        _bol_log(f"⚠ No se pudo crear ZIP: {e}", "warn")
        return None


def _bol_reset_state():
    with _bolivar_lock:
        _bolivar_state.update({
            "running": False, "stopping": False, "finished": False,
            "error": None, "logs": [],
            "stats": {"total": 0, "descargadas": 0, "errores": 0},
            "errores_detalle": [], "descargas_exitosas": [],
            "facturas_permitidas": [],
        })


def run_bolivar_automation(usuario: str, password: str,
                            periodo: str, download_path: str):
    """Automatización Playwright para descarga Bolívar SOAT desde Activa IT."""
    global _bolivar_browser, _bolivar_context, _bolivar_dl_dir
    global _bolivar_periodo, _bolivar_ips_name
    import re as _re

    dl_dir = Path(download_path)
    dl_dir.mkdir(parents=True, exist_ok=True)
    ips_nombre_actual = "IPS_SIN_NOMBRE"
    zip_parcial_generado = False
    _bolivar_dl_dir  = dl_dir
    _bolivar_periodo = periodo

    def _cerrar_traza(page):
        js = """() => {
            const headers = document.querySelectorAll('.ui-dialog-titlebar,.modal-header,[class*="header"]');
            for (const h of headers) {
                if (h.textContent && h.textContent.includes('Traza de Factura')) {
                    const dlg = h.closest('.ui-dialog,.modal,[role="dialog"]');
                    if (dlg) {
                        const btn = dlg.querySelector('.ui-dialog-titlebar-close,button.close,[aria-label*="lose"],[class*="close"]');
                        if (btn) { btn.click(); return true; }
                    }
                }
            }
            return false;
        }"""
        for fr in page.frames:
            try:
                if fr.evaluate(js):
                    time.sleep(0.5)
                    return
            except Exception:
                pass

    def _extraer_nombre_ips(page, target_frame):
        def _nit_en(frame):
            try:
                nit = frame.evaluate(
                    "() => { const m = document.body.innerText.match(/NIT\\s*:\\s*([\\d\\-\\s]+)/i); "
                    "if(m) return m[1].replace(/[^0-9]/g,''); return ''; }"
                ).strip()
                return nit
            except Exception:
                return ""
        nit = _nit_en(page) or _nit_en(target_frame)
        if not nit:
            for fr in page.frames:
                if fr not in (page, target_frame):
                    nit = _nit_en(fr)
                    if nit:
                        break
        _bol_log(f"    🔍 NIT detectado: {nit}")
        if nit and nit in _MAPA_IPS_BOLIVAR:
            return _MAPA_IPS_BOLIVAR[nit]
        js_nom = """() => {
            const kw = ["IPS","CLINICA","HOSPITAL","CENTRO","FUNDACIÓN","URGENCIAS","SALUD","ODONTOTRANS","URGETRAUMA","CORDIALIDAD"];
            for (const el of document.querySelectorAll('h1,h2,h3,h4,p,div')) {
                let t = el.innerText.trim();
                if (t.length > 5 && t.length < 100 && kw.some(k => t.toUpperCase().includes(k))) return t;
            }
            return "";
        }"""
        nombre = (page.evaluate(js_nom) or target_frame.evaluate(js_nom) or "IPS_DESCONOCIDA").strip()
        nombre = _re.sub(r'[\\/*?:"<>|]', "", nombre).strip().replace(" ", "_")
        return nombre or "IPS_DESCONOCIDA"

    def _esperar_contador(visor_page, timeout=30):
        start = time.time()
        ultimo = None
        while time.time() - start < timeout:
            if _bolivar_state.get("stopping"):
                return None
            frames = [visor_page] + list(visor_page.frames)
            for fr in frames:
                try:
                    texto = fr.evaluate("() => document.body?.innerText || ''")
                    match = re.search(r'(\d+)\s*/\s*(\d+)', texto)
                    if match:
                        val = (int(match.group(1)), int(match.group(2)))
                        if ultimo == val:
                            return val
                        ultimo = val
                        time.sleep(0.5)
                        continue
                except Exception:
                    pass
            time.sleep(0.5)
        return None

    def _download_factura(page, context, modal_frame, fac: dict,
                          dl_dir: Path, ips_nombre: str):
        import img2pdf
        from PIL import Image
        import io as _io

        num  = fac["num"]
        tipo = fac["tipo"]
        if tipo == "devolucion":
            target_label = "ActaDevolucion"
            subcarpeta   = "Devolucion"
            nombre_soporte = "ActaDevolución"
        else:
            target_label = "Envios_D"
            subcarpeta   = "Auditada"
            nombre_soporte = "Envios_D"

        ips_dir   = dl_dir / ips_nombre
        dl_subdir = ips_dir / subcarpeta
        dl_subdir.mkdir(parents=True, exist_ok=True)

        # ── Abrir traza de factura ──────────────────────────
        for intento in range(3):
            if _bolivar_state.get("stopping"):
                return
            abierto = False
            for fr in page.frames:
                try:
                    res = fr.evaluate(f"""
                        () => {{
                            const target = {json.dumps(num)};
                            for (const el of document.querySelectorAll('td,span,div,a')) {{
                                const t = (el.textContent||'').replace(/\\s+/g,' ').trim();
                                if (t.replace(/[^0-9]/g,'') === target.replace(/[^0-9]/g,'')) {{
                                    el.scrollIntoView({{block:'center'}});
                                    el.click();
                                    return true;
                                }}
                            }}
                            return false;
                        }}
                    """)
                    if res:
                        abierto = True
                        break
                except Exception:
                    pass
            if abierto:
                break
            time.sleep(2)

        # ── Esperar panel de documentos ─────────────────────
        _bol_log(f"    ⏳ Esperando panel de documentos...")
        panel_frame = None
        for _ in range(30):
            if _bolivar_state.get("stopping"):
                return
            for fr in page.frames:
                try:
                    if fr.evaluate("() => /soportes|documentos|archivos adjuntos/i.test(document.body?.innerText||'')"):
                        panel_frame = fr
                        break
                except Exception:
                    pass
            if panel_frame:
                break
            time.sleep(1)

        # ── Seleccionar documento ───────────────────────────
        archivo_seleccionado = False
        tipo_encontrado = nombre_soporte
        for intento in range(4):
            if _bolivar_state.get("stopping"):
                return
            for fr in page.frames:
                try:
                    resultado = fr.evaluate(f"""
                        () => {{
                            const label = {json.dumps(target_label)};
                            function norm(s) {{ return s.toLowerCase().normalize("NFD").replace(/[\\u0300-\\u036f]/g,""); }}
                            for (const el of document.querySelectorAll('td,div,span,li,p,tr')) {{
                                const txt = (el.innerText||'').trim();
                                if (!norm(txt).includes(norm(label))) continue;
                                let c = el.closest('div[class*="file"],li[class*="file"],tr,div[class*="item"],div[class*="attach"],div[class*="row"]') || el.closest('div,li,tr');
                                if (c) {{
                                    let chk = c.querySelector('input[type="checkbox"],input[type="radio"],[role="checkbox"]');
                                    if (!chk) chk = c.parentElement?.querySelector('input[type="checkbox"],input[type="radio"],[role="checkbox"]');
                                    if (chk) {{
                                        if (!chk.checked) {{ chk.click(); chk.checked=true; chk.dispatchEvent(new Event('change',{{bubbles:true}})); }}
                                        return {{ok:true,metodo:'checkbox'}};
                                    }}
                                    const ico = c.querySelector('img[src*="pdf"],svg[aria-label*="pdf"],i[class*="pdf"],i[class*="file"]');
                                    if (ico) {{ ico.click(); return {{ok:true,metodo:'icono_pdf'}}; }}
                                    c.click(); c.dispatchEvent(new MouseEvent('click',{{bubbles:true}}));
                                    return {{ok:true,metodo:'contenedor'}};
                                }}
                            }}
                            return {{ok:false}};
                        }}
                    """)
                    if resultado and resultado.get("ok"):
                        archivo_seleccionado = True
                        break
                except Exception:
                    pass
            if archivo_seleccionado:
                break
            time.sleep(2)

        if not archivo_seleccionado:
            raise Exception(f"No se pudo seleccionar {target_label} para factura {num}")

        # ── Abrir visor y capturar páginas ──────────────────
        _bol_log(f"    👁️ Abriendo visor documental...")
        visor_page = None
        try:
            with context.expect_page(timeout=30000) as page_info:
                for fr in page.frames:
                    try:
                        btn = fr.locator(
                            'button[title="Abrir Documento"],button[aria-label="Abrir Documento"],'
                            'button:has(i.fa-eye),button:has(i.bi-eye)'
                        ).first
                        if btn.is_visible(timeout=5000):
                            btn.click()
                            break
                    except Exception:
                        pass
            visor_page = page_info.value
            time.sleep(2)
        except Exception as e:
            raise Exception(f"No se pudo abrir el visor: {e}")

        contador = _esperar_contador(visor_page, timeout=30)
        total_paginas = contador[1] if contador else 1
        _bol_log(f"    📄 Páginas: {total_paginas}")

        imagenes = []
        for pag in range(1, total_paginas + 1):
            if _bolivar_state.get("stopping"):
                break
            screenshot = visor_page.screenshot(full_page=False)
            imagenes.append(screenshot)
            if pag < total_paginas:
                # Avanzar página
                for fr in [visor_page] + list(visor_page.frames):
                    try:
                        fr.evaluate("""() => {
                            const btns = document.querySelectorAll('button,a,[role="button"]');
                            for (const b of btns) {
                                const t = (b.textContent||b.getAttribute('aria-label')||b.title||'').toLowerCase();
                                if (t.includes('siguiente') || t.includes('next') || t === '>') { b.click(); return true; }
                            }
                            return false;
                        }""")
                        break
                    except Exception:
                        pass
                time.sleep(1)

        try:
            visor_page.close()
        except Exception:
            pass

        # Convertir capturas a PDF
        if imagenes:
            pdf_name = f"Factura_{num}_{nombre_soporte}.pdf"
            pdf_path = dl_subdir / pdf_name
            img_objs = []
            for raw in imagenes:
                img = Image.open(_io.BytesIO(raw)).convert("RGB")
                buf = _io.BytesIO()
                img.save(buf, format="JPEG", quality=90)
                img_objs.append(buf.getvalue())
            with open(pdf_path, "wb") as f_out:
                f_out.write(img2pdf.convert(img_objs))
            _bol_log(f"    ✅ PDF guardado: {pdf_path.name}")
            with _bolivar_lock:
                _bolivar_state["descargas_exitosas"].append({
                    "factura": num, "estado": fac.get("estado", ""),
                    "archivo": str(pdf_path),
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                })
        else:
            raise Exception("No se capturaron imágenes del visor")

        # Cerrar panel de traza
        _cerrar_traza(page)
        time.sleep(0.5)

    # ── Automatización principal ────────────────────────────
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            context = browser.new_context(
                accept_downloads=True, viewport={"width": 1500, "height": 900}
            )
            page = context.new_page()
            _bolivar_browser = browser
            _bolivar_context = context

            _bol_log("🔐 Iniciando sesión en Activa IT...")
            if _bolivar_state.get("stopping"):
                return
            page.goto("https://activa-it.net/Login.aspx",
                      wait_until="networkidle", timeout=60000)
            page.fill('input[placeholder="Usuario"]', usuario)
            page.fill('input[placeholder="Contraseña"]', password)
            try:
                chk = page.locator('input[type="checkbox"]').first
                if not chk.is_checked():
                    chk.check()
            except Exception:
                pass
            page.click('button:has-text("Inicio de sesión"),input[value="Inicio de sesión"]')
            page.wait_for_url("**/Index.aspx", timeout=60000)
            time.sleep(2)
            _bol_log("✅ Sesión iniciada.")

            _bol_log("📂 Navegando a módulo BI IPS...")
            time.sleep(3)

            def _find_periodo_frame():
                js = (f"() => {{ const b = (document.body?.innerText||'').toLowerCase(); "
                      f"return b.includes('{periodo.lower()}'); }}")
                for fr in page.frames:
                    try:
                        if fr.evaluate(js):
                            return fr
                    except Exception:
                        pass
                return None

            clicked = False
            for intento in range(3):
                for selector in [
                    "text=BI IPS",
                    ("text=Inteligencia de Negocio", "text=BI IPS"),
                ]:
                    try:
                        if isinstance(selector, tuple):
                            page.click(selector[0], timeout=8000)
                            time.sleep(1)
                            page.click(selector[1], timeout=8000)
                        else:
                            page.locator(selector).first.click(timeout=15000)
                        clicked = True
                        break
                    except Exception:
                        pass
                if clicked:
                    break
                time.sleep(2)
            if not clicked:
                raise Exception("No se encontró el módulo BI IPS.")

            time.sleep(3)
            target_frame = None
            for _ in range(120):
                if _bolivar_state.get("stopping"):
                    return
                target_frame = _find_periodo_frame()
                if target_frame:
                    break
                time.sleep(0.5)
            if not target_frame:
                raise Exception(f"No se localizó el período '{periodo}'.")

            ips_nombre_actual = _extraer_nombre_ips(page, target_frame)
            _bolivar_ips_name  = ips_nombre_actual
            _bol_log(f"🏥 IPS: {ips_nombre_actual}")

            # Click en columna Cant
            click_result = target_frame.evaluate(f"""
                () => {{
                    for (const row of document.querySelectorAll('tr')) {{
                        const cells = row.querySelectorAll('td');
                        if (cells.length < 3) continue;
                        if (cells[0].textContent.trim() !== '{periodo}') continue;
                        const links = row.querySelectorAll('a');
                        if (!links.length) return {{ok:false,reason:'sin_links'}};
                        const fl = links[0];
                        if (fl.textContent.trim() === '0') return {{ok:false,reason:'cant_cero'}};
                        fl.scrollIntoView({{block:'center'}}); fl.click();
                        return {{ok:true,value:fl.textContent.trim()}};
                    }}
                    return {{ok:false,reason:'fila_no_encontrada'}};
                }}
            """)
            if click_result.get("reason") == "cant_cero":
                _bol_log(f"ℹ️ El período '{periodo}' tiene 0 facturas.", "warn")
                browser.close()
                return
            if not click_result.get("ok"):
                raise Exception(f"No se hizo click en Cant: {click_result.get('reason')}")

            # Esperar modal
            modal_frame = None
            for _ in range(60):
                if _bolivar_state.get("stopping"):
                    return
                for fr in page.frames:
                    try:
                        if fr.evaluate("() => /Listado de facturas recibidas/i.test(document.body?.innerText||'')"):
                            modal_frame = fr
                            break
                    except Exception:
                        pass
                if modal_frame:
                    break
                time.sleep(0.5)
            if not modal_frame:
                raise Exception("Modal 'Listado de facturas recibidas' no apareció.")

            # Esperar datos
            data_frame = None
            for _ in range(120):
                if _bolivar_state.get("stopping"):
                    return
                for fr in page.frames:
                    try:
                        if fr.evaluate("() => /Pendiente de recibir Informaci|Devoluci[oó]n de entrada/i.test(document.body?.innerText||'')"):
                            data_frame = fr
                            break
                    except Exception:
                        pass
                if data_frame:
                    break
                time.sleep(0.5)
            if not data_frame:
                _bol_log("⚠ No se encontraron facturas con los estados objetivo.", "warn")
                browser.close()
                return

            time.sleep(2)

            # ── Extracción de facturas ──────────────────────
            js_extract = r"""
            (state) => {
                const ESTADOS = [
                    { regex: /auditada\s*:\s*pendiente\s+de\s+recibir\s+informaci[oó]n/i, tipo: 'auditada',
                      nombre: 'Auditada: Pendiente de recibir Informacion' },
                    { regex: /en\s+radicaci[oó]n\s*:\s*devoluci[oó]n\s+de\s+entrada/i, tipo: 'devolucion',
                      nombre: 'En radicacion: Devolución de entrada' },
                    { regex: /en\s+auditori?a\s*:\s*pendiente\s+de\s+informar\s+orden\s+de\s+pago\s+al\s+pagador/i, tipo: 'auditada',
                      nombre: 'En auditoria: Pendiente de informar Orden de pago al Pagador' },
                ];
                function norm(s) {
                    return s.toLowerCase().normalize("NFD").replace(/[\u0300-\u036f]/g,"").replace(/[''´`]/g,"'");
                }
                const nuevas = [];
                for (const fila of document.querySelectorAll('tr,[role="row"],li')) {
                    const fullText = (fila.innerText||'').replace(/\s+/g,' ').trim();
                    if (!fullText || fullText.length < 20 || fullText.length > 400) continue;
                    const textoNorm = norm(fullText);
                    let tipoDetectado = null, nombreEstado = null;
                    for (const e of ESTADOS) {
                        if (e.regex.test(textoNorm)) { tipoDetectado=e.tipo; nombreEstado=e.nombre; break; }
                    }
                    if (!tipoDetectado) continue;
                    let numFactura = null;
                    for (const token of fullText.split(/\s+/)) {
                        const digits = token.replace(/^#+/,'').replace(/\D/g,'');
                        if (digits.length >= 6 && digits.length <= 12) { numFactura=digits; break; }
                    }
                    if (!numFactura || state.seen.includes(numFactura)) continue;
                    const botId = 'bot_' + state.nextId++;
                    fila.setAttribute('data-bot-row-id', botId);
                    nuevas.push({ botId, num: numFactura, tipo: tipoDetectado, estado: nombreEstado });
                    state.seen.push(numFactura);
                }
                return { nuevas, total: state.seen.length };
            }
            """
            extract_state = {"nextId": 0, "seen": []}
            facturas_acumuladas = []
            rondas_sin_nuevos = 0
            for ronda in range(20):
                if _bolivar_state.get("stopping"):
                    return
                try:
                    res = data_frame.evaluate(js_extract, extract_state)
                except Exception as e:
                    _bol_log(f"  ⚠ Error extracción ronda {ronda+1}: {e}", "warn")
                    res = {"nuevas": []}
                nuevas = res.get("nuevas", [])
                if nuevas:
                    facturas_acumuladas.extend(nuevas)
                    rondas_sin_nuevos = 0
                    _bol_log(f"  Ronda {ronda+1}: +{len(nuevas)} (Total: {len(facturas_acumuladas)})")
                else:
                    rondas_sin_nuevos += 1
                extract_state["seen"] = list(set(extract_state["seen"] + [n["num"] for n in nuevas]))
                if rondas_sin_nuevos >= 5:
                    break
                try:
                    data_frame.evaluate("() => { const s=document.querySelectorAll('div,table,tbody,[class*=\"scroll\"]'); for(const e of s) if(e.scrollHeight>e.clientHeight+20) e.scrollTop+=e.clientHeight*0.8; window.scrollBy(0,window.innerHeight*0.8); }")
                except Exception:
                    pass
                time.sleep(0.5)

            _bol_log(f"📊 {len(facturas_acumuladas)} facturas detectadas.")

            # ── Filtros: progreso y lista permitida ─────────
            ips_dir    = dl_dir / ips_nombre_actual
            completadas = _bol_cargar_progreso(ips_dir)
            facturas_pendientes = []
            for fac in facturas_acumuladas:
                if fac["num"] in completadas:
                    _bol_log(f"⏭️ {fac['num']} ya descargada.")
                    with _bolivar_lock:
                        _bolivar_state["stats"]["descargadas"] += 1
                else:
                    facturas_pendientes.append(fac)

            with _bolivar_lock:
                permitidas = list(_bolivar_state.get("facturas_permitidas", []))
            if permitidas:
                orig = len(facturas_pendientes)
                facturas_pendientes = [f for f in facturas_pendientes if f["num"] in permitidas]
                _bol_log(f"📋 Filtro activo: {len(facturas_pendientes)} de {orig}")

            with _bolivar_lock:
                _bolivar_state["stats"]["total"] = (
                    len(facturas_pendientes) + _bolivar_state["stats"]["descargadas"]
                )

            _bol_log(f"📋 Pendientes: {len(facturas_pendientes)}")
            if not facturas_pendientes:
                _bol_log("ℹ️ No hay facturas pendientes.")
                browser.close()
                with _bolivar_lock:
                    ex = list(_bolivar_state["descargas_exitosas"])
                    er = list(_bolivar_state["errores_detalle"])
                _bol_generar_excel(dl_dir, periodo, ips_nombre_actual, ex, er)
                _bol_crear_zip(dl_dir, periodo, ips_nombre_actual)
                return

            # ── Procesar facturas ───────────────────────────
            for idx, fac in enumerate(facturas_pendientes, 1):
                if _bolivar_state.get("stopping"):
                    _bol_log("🛑 Proceso detenido por el usuario.")
                    if not zip_parcial_generado:
                        _bol_crear_zip(dl_dir, periodo, ips_nombre_actual)
                        zip_parcial_generado = True
                    return
                _bol_log(f"[{idx}/{len(facturas_pendientes)}] Factura {fac['num']} ({fac['tipo']})...")
                try:
                    _download_factura(page, context, data_frame, fac,
                                      dl_dir, ips_nombre_actual)
                    with _bolivar_lock:
                        _bolivar_state["stats"]["descargadas"] += 1
                    completadas.add(fac["num"])
                    _bol_guardar_progreso(ips_dir, completadas)
                    _bol_log(f"  ✅ {fac['num']}", "success")
                except Exception as e:
                    with _bolivar_lock:
                        _bolivar_state["stats"]["errores"] += 1
                        err_info = {
                            "factura": fac["num"], "estado": fac.get("estado", ""),
                            "error": str(e), "captura": "",
                            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        }
                        try:
                            errores_dir = ips_dir / "Errores"
                            errores_dir.mkdir(parents=True, exist_ok=True)
                            cap_path = errores_dir / f"ERROR_{fac['num']}.png"
                            page.screenshot(path=str(cap_path))
                            err_info["captura"] = str(cap_path)
                        except Exception:
                            pass
                        _bolivar_state["errores_detalle"].append(err_info)
                    _bol_log(f"  ⚠ {e}", "error")
                    _cerrar_traza(page)
                    time.sleep(1)

            browser.close()
            with _bolivar_lock:
                ex = list(_bolivar_state["descargas_exitosas"])
                er = list(_bolivar_state["errores_detalle"])
            _bol_generar_excel(dl_dir, periodo, ips_nombre_actual, ex, er)
            _bol_crear_zip(dl_dir, periodo, ips_nombre_actual)
            _bol_log("🎉 Proceso completado.", "success")

    except Exception as e:
        if not _bolivar_state.get("stopping"):
            _bol_log(f"💥 Error crítico: {e}", "error")
            with _bolivar_lock:
                _bolivar_state["error"] = str(e)
        else:
            _bol_log("Proceso detenido por el usuario.")
        if not zip_parcial_generado and _bolivar_dl_dir and _bolivar_periodo and _bolivar_ips_name:
            _bol_crear_zip(_bolivar_dl_dir, _bolivar_periodo, _bolivar_ips_name)
    finally:
        with _bolivar_lock:
            _bolivar_state["running"]  = False
            _bolivar_state["finished"] = True
            _bolivar_state["stopping"] = False
        _bolivar_browser = _bolivar_context = None
        _bolivar_dl_dir  = _bolivar_periodo = _bolivar_ips_name = None


# ══════════════════════════════════════════════════════════════
#  RUTAS API
# ══════════════════════════════════════════════════════════════

@app.route("/api/health")
def health():
    desapilar_ok  = _find_desapilar_py() is not None
    descarga_ok   = _find_descarga_py() is not None
    unificar_ok   = _find_unificar_py() is not None
    renombrar_ok  = _find_renombrar_py() is not None
    bolivar_ok    = _find_bolivar_py() is not None
    portal_ok     = _find_portal_py() is not None
    matriz_ok     = any((d / "MATRIZ_IPS.xlsx").exists()
                        for d in [BASE_DIR, BASE_DIR.parent])
    return jsonify({
        "status":        "ok",
        "desapilar_py":  desapilar_ok,
        "descarga_py":   descarga_ok,
        "unificar_py":   unificar_ok,
        "renombrar_py":  renombrar_ok,
        "bolivar_py":    bolivar_ok,
        "portal_py":     portal_ok,
        "matriz_ips":    matriz_ok,
    })


@app.route("/api/stream/<session_id>")
def stream_logs(session_id):
    """SSE endpoint para logs en tiempo real."""
    def generate():
        q = get_queue(session_id)
        while True:
            try:
                item = q.get(timeout=30)
                yield f"data: {json.dumps(item)}\n\n"
                if item.get("type") == "done":
                    break
            except queue.Empty:
                yield "data: {\"type\":\"ping\"}\n\n"
    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control":               "no-cache",
            "X-Accel-Buffering":           "no",
            "Access-Control-Allow-Origin": "*",
        },
    )


@app.route("/api/desapilar", methods=["POST"])
def api_desapilar():
    """Recibe PDFs, los desapila y devuelve ZIP con resultados."""
    session_id = request.form.get("session_id", f"ds_{int(time.time())}")
    files      = request.files.getlist("pdfs")

    if not files:
        return jsonify({"ok": False, "error": "No se recibieron archivos"}), 400

    # Guardar uploads
    saved = []
    for f in files:
        if f.filename.lower().endswith(".pdf"):
            fn   = secure_filename(f.filename)
            dest = UPLOAD_DIR / f"{session_id}_{fn}"
            f.save(str(dest))
            saved.append(dest)

    if not saved:
        return jsonify({"ok": False, "error": "Ningún archivo PDF válido"}), 400

    def run():
        try:
            result = desapilar_standalone(saved, session_id)
            push_done(session_id, result)
        except Exception as e:
            push_log(session_id, f"Error crítico: {e}", "error")
            push_done(session_id, {"ok": False, "error": str(e)})
        finally:
            for p in saved:
                try: p.unlink()
                except: pass

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"ok": True, "session_id": session_id})


@app.route("/api/descarga", methods=["POST"])
def api_descarga():
    """Inicia descarga + desapilar desde correo."""
    data       = request.get_json() or {}
    session_id = data.get("session_id", f"dc_{int(time.time())}")

    def run():
        try:
            result = run_descarga_y_desapilar(data, session_id)
            push_done(session_id, result)
        except Exception as e:
            push_log(session_id, f"Error crítico: {e}", "error")
            push_done(session_id, {"ok": False, "error": str(e)})

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"ok": True, "session_id": session_id})


@app.route("/api/unificar", methods=["POST"])
def api_unificar():
    """Recibe PDFs, unifica los que comparten base numérica y devuelve ZIP."""
    session_id = request.form.get("session_id", f"uf_{int(time.time())}")
    files      = request.files.getlist("pdfs")

    if not files:
        return jsonify({"ok": False, "error": "No se recibieron archivos"}), 400

    saved = []
    for f in files:
        if f.filename.lower().endswith(".pdf"):
            fn   = secure_filename(f.filename)
            dest = UPLOAD_DIR / f"{session_id}_{fn}"
            f.save(str(dest))
            saved.append(dest)

    if not saved:
        return jsonify({"ok": False, "error": "Ningún archivo PDF válido"}), 400

    def run():
        try:
            result = run_unificar(saved, session_id)
            push_done(session_id, result)
        except Exception as e:
            push_log(session_id, f"Error crítico: {e}", "error")
            push_done(session_id, {"ok": False, "error": str(e)})
        finally:
            for p in saved:
                try: p.unlink()
                except: pass

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"ok": True, "session_id": session_id})


@app.route("/api/download/<path:rel_path>")
def download_file(rel_path):
    abs_path = BASE_DIR / rel_path
    if not abs_path.exists():
        return jsonify({"error": "Archivo no encontrado"}), 404
    return send_file(str(abs_path), as_attachment=True)


@app.route("/api/renombrar", methods=["POST"])
def api_renombrar():
    """Recibe PDFs, los renombra según aseguradora y devuelve ZIP + reporte."""
    session_id = request.form.get("session_id", f"rn_{int(time.time())}")
    files      = request.files.getlist("pdfs")

    if not files:
        return jsonify({"ok": False, "error": "No se recibieron archivos"}), 400

    saved = []
    for f in files:
        if f.filename.lower().endswith(".pdf"):
            fn   = secure_filename(f.filename)
            dest = UPLOAD_DIR / f"{session_id}_{fn}"
            f.save(str(dest))
            saved.append(dest)

    if not saved:
        return jsonify({"ok": False, "error": "Ningún archivo PDF válido"}), 400

    def run():
        try:
            result = run_renombrar(saved, session_id)
            push_done(session_id, result)
        except Exception as e:
            push_log(session_id, f"Error crítico: {e}", "error")
            push_done(session_id, {"ok": False, "error": str(e)})
        finally:
            for p in saved:
                try: p.unlink()
                except: pass

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"ok": True, "session_id": session_id})


@app.route("/api/config/check", methods=["POST"])
def check_config():
    """Verifica archivos de configuracion separados por modulo."""
    def exists(name):
        return any((d / name).exists() for d in [BASE_DIR, BASE_DIR.parent])

    descarga = {
        "DESCARGA_GLOSAS.py":  exists("DESCARGA_GLOSAS.py"),
        "DESAPILAR_GLOSAS.py": exists("DESAPILAR_GLOSAS.py"),
        "MATRIZ_IPS.xlsx":     exists("MATRIZ_IPS.xlsx"),
    }
    desapilar = {
        "DESAPILAR_GLOSAS.py": exists("DESAPILAR_GLOSAS.py"),
        "MATRIZ_IPS.xlsx":     exists("MATRIZ_IPS.xlsx"),
    }
    unificar = {
        "unificar_pdfs.py": exists("unificar_pdfs.py"),
    }
    renombrar = {
        "renombrar_auditoria.py": exists("renombrar_auditoria.py"),
    }
    bolivar = {
        "DESCARGA_BOLIVAR.py": exists("DESCARGA_BOLIVAR.py"),
        "Playwright": True,  # Se verifica en runtime
    }
    portal = {
        "DESCARGA_PORTAL.py": exists("DESCARGA_PORTAL.py"),
        "Playwright": True,
    }
    return jsonify({"descarga": descarga, "desapilar": desapilar,
                    "unificar": unificar, "renombrar": renombrar,
                    "bolivar": bolivar, "portal": portal})


@app.route("/")
@app.route("/index.html")
def index():
    for search in [BASE_DIR, BASE_DIR.parent, BASE_DIR.parent.parent]:
        html = search / "static" / "index.html"
        if html.exists():
            return send_file(str(html))
    return "index.html no encontrado. Verifique que exista la carpeta static/", 404




# ── Rutas Portal Activa IT (Previsora SOAT) ──────────────────

def _load_portal_mod():
    """
    Carga DESCARGA_PORTAL.py dinámicamente una sola vez y conserva el módulo
    (con su estado global) para todas las llamadas posteriores.
    """
    import importlib.util

    # Si ya está cargado y con estado válido, reusar
    cached = sys.modules.get("portal_mod")
    if cached is not None and hasattr(cached, "_job_state"):
        return cached

    portal_py = _find_portal_py()
    if not portal_py:
        return None

    spec = importlib.util.spec_from_file_location("portal_mod", str(portal_py))
    mod  = importlib.util.module_from_spec(spec)

    # Parchear __file__ ANTES de exec para que BASE_DIR/DOWNLOAD_DIR resuelvan bien
    mod.__file__ = str(portal_py)
    mod.__spec__ = spec

    # Ejecutar primero; registrar en sys.modules despues (evita modulo vacio si falla)
    spec.loader.exec_module(mod)
    sys.modules["portal_mod"] = mod
    return mod


@app.route("/api/portal/start", methods=["POST", "OPTIONS"])
def portal_start():
    """Inicia la descarga automatica desde Activa IT (Previsora SOAT)."""
    if request.method == "OPTIONS":
        return jsonify({"ok": True}), 200
    try:
        data        = request.get_json() or {}
        usuario     = data.get("usuario", "").strip()
        password    = data.get("password", "").strip()
        periodo_raw = data.get("periodo", "").strip()

        if not all([usuario, password, periodo_raw]):
            return jsonify({"ok": False, "error": "Faltan campos requeridos"}), 400

        mod = _load_portal_mod()
        if not mod:
            return jsonify({"ok": False, "error": "DESCARGA_PORTAL.py no encontrado en la carpeta del proyecto"}), 500

        # Convertir YYYY-MM a MmmYY
        MESES_P = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
        periodo = periodo_raw
        m1 = re.match(r'^(\d{4})-(\d{1,2})$', periodo_raw)
        if m1:
            mes_idx = int(m1.group(2)) - 1
            if 0 <= mes_idx < 12:
                periodo = MESES_P[mes_idx] + m1.group(1)[2:]
        else:
            m2 = re.match(r'^([a-zA-Z]{3})(\d{2,4})$', periodo_raw)
            if m2:
                periodo = m2.group(1).capitalize() + m2.group(2)[-2:]

        with mod._job_lock:
            if mod._job_state["running"]:
                return jsonify({"ok": False, "error": "Ya hay un proceso en ejecucion"}), 409
            mod._job_state.update({
                "running": True, "finished": False, "error": None,
                "stats": {"total": 0, "descargadas": 0, "errores": 0},
                "logs": [], "errores_detalle": [], "descargas_exitosas": [],
                "zip_path": None, "excel_path": None,
            })

        dl_path = str(mod.DOWNLOAD_DIR / periodo)
        threading.Thread(
            target=mod.run_portal,
            args=(usuario, password, periodo, dl_path),
            daemon=True,
        ).start()
        return jsonify({"ok": True, "periodo": periodo, "download_path": dl_path})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error interno: {e}"}), 500


@app.route("/api/portal/diagnose")
def portal_diagnose():
    """Verifica el estado del modulo portal para depuracion."""
    try:
        portal_py = _find_portal_py()
        if not portal_py:
            return jsonify({"ok": False, "error": "DESCARGA_PORTAL.py no encontrado",
                           "buscado_en": [str(BASE_DIR), str(BASE_DIR.parent)]})
        mod = _load_portal_mod()
        if not mod:
            return jsonify({"ok": False, "error": "Modulo cargado como None"})
        return jsonify({
            "ok": True,
            "portal_py": str(portal_py),
            "tiene_run_portal": hasattr(mod, "run_portal"),
            "tiene_job_state": hasattr(mod, "_job_state"),
            "tiene_job_lock": hasattr(mod, "_job_lock"),
            "download_dir": str(getattr(mod, "DOWNLOAD_DIR", "?")),
            "playwright_ok": getattr(mod, "PLAYWRIGHT_OK", False),
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "tipo": type(e).__name__})


@app.route("/api/portal/stop", methods=["POST"])
def portal_stop():
    try:
        mod = _load_portal_mod()
        if not mod:
            return jsonify({"ok": False, "error": "DESCARGA_PORTAL.py no encontrado"}), 500
        with mod._job_lock:
            if not mod._job_state["running"]:
                return jsonify({"ok": False, "message": "No hay proceso en ejecucion"}), 400
        mod.stop_portal()
        return jsonify({"ok": True, "message": "Deteniendo proceso..."})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/portal/status")
def portal_status():
    try:
        mod = _load_portal_mod()
        if not mod:
            return jsonify({"ok": False, "error": "DESCARGA_PORTAL.py no encontrado"}), 500
        since = int(request.args.get("since", 0))
        with mod._job_lock:
            logs  = list(mod._job_state["logs"][since:])
            stats = dict(mod._job_state["stats"])
            resp  = {
                "ok":       True,
                "running":  mod._job_state["running"],
                "finished": mod._job_state["finished"],
                "error":    mod._job_state["error"],
                "stats":    stats,
                "logs":     logs,
            }
            if mod._job_state["zip_path"]:
                zp = Path(mod._job_state["zip_path"])
                try:
                    resp["zip_path"] = str(zp.relative_to(BASE_DIR))
                except ValueError:
                    resp["zip_path"] = mod._job_state["zip_path"]
            if mod._job_state["excel_path"]:
                ep = Path(mod._job_state["excel_path"])
                try:
                    resp["excel_path"] = str(ep.relative_to(BASE_DIR))
                except ValueError:
                    resp["excel_path"] = mod._job_state["excel_path"]
        return jsonify(resp)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/portal/reset", methods=["POST"])
def portal_reset():
    try:
        mod = _load_portal_mod()
        if not mod:
            return jsonify({"ok": False, "error": "DESCARGA_PORTAL.py no encontrado"}), 500
        data    = request.get_json() or {}
        periodo = data.get("periodo", "").strip()
        if periodo:
            mod.reset_progreso(periodo)
        mod.reset_state()
        return jsonify({"ok": True, "message": "Estado reiniciado."})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/portal/upload_facturas", methods=["POST"])
def portal_upload_facturas():
    """Carga lista de facturas permitidas para el módulo portal (CSV o Excel)."""
    import csv as _csv
    from io import BytesIO as _BytesIO

    mod = _load_portal_mod()
    if not mod:
        return jsonify({"ok": False, "error": "DESCARGA_PORTAL.py no encontrado"}), 500

    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No se envió ningún archivo"}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"ok": False, "error": "Archivo vacío"}), 400
    try:
        fname    = file.filename.lower()
        facturas = []
        if fname.endswith(".csv"):
            content = file.read().decode("utf-8")
            reader  = _csv.DictReader(content.splitlines())
            for row in reader:
                for col, val in row.items():
                    if "factura" in col.lower():
                        facturas.append(val.strip())
                        break
        elif fname.endswith((".xls", ".xlsx")):
            try:
                import openpyxl as _ox
            except ImportError:
                return jsonify({"ok": False, "error": "openpyxl no instalado"}), 500
            wb      = _ox.load_workbook(_BytesIO(file.read()), data_only=True)
            ws      = wb.active
            col_idx = None
            for cell in ws[1]:
                if cell.value and "factura" in str(cell.value).lower():
                    col_idx = cell.column
                    break
            if col_idx is None:
                return jsonify({"ok": False, "error": "No se encontró columna 'factura'"}), 400
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[col_idx - 1]
                if val:
                    facturas.append(str(val).strip())
        else:
            return jsonify({"ok": False, "error": "Formato no soportado. Use CSV o Excel"}), 400

        limpias = [re.sub(r'\D', '', f) for f in facturas if re.sub(r'\D', '', f)]
        if not limpias:
            return jsonify({"ok": False, "error": "No se encontraron números válidos"}), 400
        mod.set_facturas_permitidas(limpias)
        mod._log(f"📄 {len(limpias)} facturas cargadas como filtro.")
        return jsonify({"ok": True, "count": len(limpias), "facturas": limpias[:10]})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error: {e}"}), 500


# ── Rutas Bolívar SOAT ──────────────────────────────────────

@app.route("/api/bolivar/start", methods=["POST"])
def bolivar_start():
    """Inicia la descarga automática desde Activa IT (Bolívar SOAT)."""
    data     = request.get_json() or {}
    usuario  = data.get("usuario", "").strip()
    password = data.get("password", "").strip()
    periodo_raw = data.get("periodo", "").strip()

    if not all([usuario, password, periodo_raw]):
        return jsonify({"ok": False, "error": "Faltan campos requeridos"}), 400

    periodos = _bol_parse_periodo(periodo_raw)
    if not periodos:
        return jsonify({
            "ok": False,
            "error": f"Formato de período inválido: '{periodo_raw}'. Use MmmYY (ej: May26) o rango MmmYY-MmmYY",
        }), 400

    with _bolivar_lock:
        if _bolivar_state["running"]:
            return jsonify({"ok": False, "error": "Ya hay un proceso en ejecución"}), 409
        _bolivar_state.update({
            "running": True, "finished": False, "error": None,
            "stats": {"total": 0, "descargadas": 0, "errores": 0},
            "errores_detalle": [], "descargas_exitosas": [],
        })

    periodo    = periodos[0]
    dl_path    = str(BOLIVAR_DOWNLOAD_DIR / periodo_raw)
    threading.Thread(
        target=run_bolivar_automation,
        args=(usuario, password, periodo, dl_path),
        daemon=True,
    ).start()
    return jsonify({"ok": True, "periodo": periodo, "download_path": dl_path})


@app.route("/api/bolivar/stop", methods=["POST"])
def bolivar_stop():
    with _bolivar_lock:
        if not _bolivar_state["running"]:
            return jsonify({"ok": False, "message": "No hay proceso en ejecución"}), 400
        _bolivar_state["stopping"] = True
    _bol_log("🛑 Detención solicitada por el usuario.", "warn")
    if _bolivar_browser:
        try:
            _bolivar_browser.close()
        except Exception:
            pass
    return jsonify({"ok": True, "message": "Deteniendo proceso..."})


@app.route("/api/bolivar/status")
def bolivar_status():
    since = int(request.args.get("since", 0))
    with _bolivar_lock:
        logs   = list(_bolivar_state["logs"][since:])
        stats  = dict(_bolivar_state["stats"])
        estado = {
            "ok":       True,
            "running":  _bolivar_state["running"],
            "finished": _bolivar_state["finished"],
            "error":    _bolivar_state["error"],
            "stats":    stats,
            "logs":     logs,
        }
        # Incluir paths de ZIP y Excel si terminó
        if _bolivar_state["finished"] and not _bolivar_state["error"]:
            # Buscar último ZIP generado
            if BOLIVAR_DOWNLOAD_DIR.exists():
                zips = sorted(BOLIVAR_DOWNLOAD_DIR.rglob("facturas_*.zip"),
                              key=lambda p: p.stat().st_mtime, reverse=True)
                if zips:
                    try:
                        estado["zip_path"] = str(zips[0].relative_to(BASE_DIR))
                    except ValueError:
                        estado["zip_path"] = str(zips[0])
                excels = sorted(BOLIVAR_DOWNLOAD_DIR.rglob("reporte_*.xlsx"),
                                key=lambda p: p.stat().st_mtime, reverse=True)
                if excels:
                    try:
                        estado["excel_path"] = str(excels[0].relative_to(BASE_DIR))
                    except ValueError:
                        estado["excel_path"] = str(excels[0])
    return jsonify(estado)


@app.route("/api/bolivar/reset", methods=["POST"])
def bolivar_reset():
    data    = request.get_json() or {}
    periodo = data.get("periodo", "").strip()
    with _bolivar_lock:
        if _bolivar_state["running"] and _bolivar_browser:
            try:
                _bolivar_browser.close()
            except Exception:
                pass
    if periodo:
        periodo_dir = BOLIVAR_DOWNLOAD_DIR / periodo
        if periodo_dir.exists():
            for prog in periodo_dir.glob("*/progreso.json"):
                try:
                    prog.unlink()
                except Exception:
                    pass
    _bol_reset_state()
    return jsonify({"ok": True, "message": "Estado reiniciado."})


@app.route("/api/bolivar/upload_facturas", methods=["POST"])
def bolivar_upload_facturas():
    """Carga una lista de facturas permitidas (CSV o Excel)."""
    import csv as _csv
    from io import BytesIO as _BytesIO

    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No se envió ningún archivo"}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"ok": False, "error": "Archivo vacío"}), 400
    try:
        fname = file.filename.lower()
        facturas = []
        if fname.endswith(".csv"):
            content = file.read().decode("utf-8")
            reader  = _csv.DictReader(content.splitlines())
            for row in reader:
                for col, val in row.items():
                    if "factura" in col.lower():
                        facturas.append(val.strip())
                        break
        elif fname.endswith((".xls", ".xlsx")):
            if not _EXCEL_OK:
                return jsonify({"ok": False, "error": "openpyxl no instalado"}), 500
            wb = _openpyxl.load_workbook(_BytesIO(file.read()), data_only=True)
            ws = wb.active
            col_idx = None
            for cell in ws[1]:
                if cell.value and "factura" in str(cell.value).lower():
                    col_idx = cell.column
                    break
            if col_idx is None:
                return jsonify({"ok": False, "error": "No se encontró columna 'factura'"}), 400
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[col_idx - 1]
                if val:
                    facturas.append(str(val).strip())
        else:
            return jsonify({"ok": False, "error": "Formato no soportado. Use CSV o Excel"}), 400

        limpias = [re.sub(r'\D', '', f) for f in facturas if re.sub(r'\D', '', f)]
        if not limpias:
            return jsonify({"ok": False, "error": "No se encontraron números válidos"}), 400
        with _bolivar_lock:
            _bolivar_state["facturas_permitidas"] = limpias
        _bol_log(f"📄 {len(limpias)} facturas cargadas como filtro.")
        return jsonify({"ok": True, "count": len(limpias), "facturas": limpias[:10]})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error: {e}"}), 500


@app.route("/api/bolivar/progreso")
def bolivar_progreso():
    periodo = request.args.get("periodo", "")
    if not periodo:
        return jsonify({"ok": False, "error": "Se requiere 'periodo'"}), 400
    periodo_dir = BOLIVAR_DOWNLOAD_DIR / periodo
    if not periodo_dir.exists():
        return jsonify({"ok": True, "completadas": [], "mensaje": "Sin datos para este período"})
    prog_files = list(periodo_dir.glob("*/progreso.json"))
    if not prog_files:
        return jsonify({"ok": True, "completadas": [], "mensaje": "Sin progreso registrado"})
    try:
        data = json.loads(prog_files[0].read_text(encoding="utf-8"))
        completadas = data.get("completadas", [])
        return jsonify({"ok": True, "completadas": completadas,
                        "cantidad": len(completadas), "ips": prog_files[0].parent.name,
                        "actualizado": data.get("actualizado", "")})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
# ══════════════════════════════════════════════════════════════
#  RUTAS PORTAL ACTIVA IT
# ══════════════════════════════════════════════════════════════

@app.route("/portal/start", methods=["POST", "OPTIONS"])
def portal_start():
    if request.method == "OPTIONS":
        return jsonify({"ok": True})
    data = request.get_json() or {}
    usuario  = data.get("usuario", "")
    password = data.get("password", "")
    periodo  = data.get("periodo", "")
    if not usuario or not password or not periodo:
        return jsonify({"ok": False, "error": "Faltan campos: usuario, password, periodo"}), 400
    try:
        from DESCARGA_PORTAL import run_portal, reset_state, _job_state, _job_lock
        with _job_lock:
            if _job_state["running"]:
                return jsonify({"ok": False, "error": "Ya hay un proceso en curso"}), 409
        reset_state()
        threading.Thread(target=run_portal, args=(usuario, password, periodo), daemon=True).start()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/portal/stop", methods=["POST"])
def portal_stop():
    try:
        from DESCARGA_PORTAL import stop_portal
        stop_portal()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/portal/status")
def portal_status():
    try:
        since = int(request.args.get("since", 0))
        from DESCARGA_PORTAL import get_state, get_logs_since
        state = get_state()
        logs  = get_logs_since(since)
        return jsonify({
            "ok":       True,
            "running":  state["running"],
            "finished": state["finished"],
            "error":    state["error"],
            "stats":    state["stats"],
            "logs":     logs,
            "zip_path": state["zip_path"],
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/portal/reset", methods=["POST"])
def portal_reset():
    try:
        from DESCARGA_PORTAL import reset_state
        reset_state()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

if __name__ == "__main__":
    print("=" * 55)
    print("  SISTEMA GLOSAS - APP WEB LOCAL")
    print("  http://localhost:5000")
    print("=" * 55)
    app.run(debug=False, host="0.0.0.0", port=5000, threaded=True)
