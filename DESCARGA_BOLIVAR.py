"""
Activa IT - Descargador automático de cartas glosa (Bolívar SOAT)
Versión con espera explícita del contador de páginas (ej. "1/11").
"""

import os
import re
import json
import csv
import time
import threading
import logging
import zipfile
import shutil
from datetime import datetime
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_from_directory, send_file
from io import BytesIO

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ openpyxl no instalado. No se generará el archivo Excel.")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

BASE_DIR = Path(__file__).parent
DOWNLOAD_DIR = BASE_DIR / "downloads"
DOWNLOAD_DIR.mkdir(exist_ok=True)

# ==================== MAPA DE IPS POR NIT ====================
MAPA_IPS = {
    "900267064": "INVERSIONES_AZALUD_CLINICA_BAHIA",
    "900827065": "CENTRO_DE_DIAGNOSTICO_E_IMAGENES_BAHIA",
    "900657731": "CENTRO_MEDICO_Y_DE_REHABILITACION_BAHIA",
    "900826509": "RED_DE_URGENCIAS_DEL_MAGDALENA",
    "900513306": "FUNDACION_MARIA_REINA",
    "900600550": "INVERSIONES_MEDICAS_BARU",
    "900954800": "CENTRO_MEDICO_Y_DE_REHABILITACION_BARU",
    "900631361": "INVERSIONES_MEDICAS_VALLESALUD",
    "900257333": "ODONTOTRANS",
    "901081281": "URGETRAUMA",
    "900792417": "RED_DE_URGENCIAS_DE_LA_COSTA_PACIFICA",
    "901959993": "CLINICA_CORDIALIDAD",
    "900002780": "FUNDACION_CAMPBELL",
    "901523868": "MOVID_IPS_SAS",
    "901057487": "TECNOLOGIA_DIAGNOSTICA_DEL_VALLE",
    "900558595": "FUNDACION_MEDICA_CAMPBELL",
    "901149757": "UNIDAD_MEDICA_DE_TRAUMA_VALLE_SALUD",
    "900900754": "CLINICA_VALLE_SALUD_SAN_FERNANDO",
    "900469882": "CENTRO_MEDICO_SERVISALUD_INTEGRAL_IPS_SAS",
    "802024329": "RED_DE_URGENCIA_DE_LA_COSTA_LTDA",
    "900847382": "CENTRO_MEDICO_Y_DE_REHABILITACION_VALLE_SALUD",
}

# ==================== ESTADO GLOBAL ====================
job_state = {
    "running": False,
    "stopping": False,
    "logs": [],
    "stats": {"total": 0, "descargadas": 0, "errores": 0},
    "finished": False,
    "error": None,
    "errores_detalle": [],
    "descargas_exitosas": [],
    "facturas_permitidas": [],
}
job_lock = threading.Lock()
current_browser = None
current_context = None
current_dl_dir = None
current_periodo = None
current_ips_nombre = None

# ==================== UTILIDADES DE PERÍODOS ====================
MESES = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]

def validar_periodo(p):
    if not p or len(p) < 5:
        return False
    mes = p[:3]
    anio = p[3:]
    return mes in MESES and re.match(r'^\d{2}$', anio)

def generar_rango_periodos(inicio, fin):
    if not validar_periodo(inicio) or not validar_periodo(fin):
        return []
    mes_inicio = MESES.index(inicio[:3])
    anio_inicio = int(inicio[3:])
    mes_fin = MESES.index(fin[:3])
    anio_fin = int(fin[3:])
    fecha_inicio = anio_inicio * 100 + mes_inicio
    fecha_fin = anio_fin * 100 + mes_fin
    if fecha_fin < fecha_inicio:
        return []
    periodos = []
    anio = anio_inicio
    mes = mes_inicio
    while True:
        anio_str = str(anio).zfill(2)
        periodos.append(MESES[mes] + anio_str)
        if anio == anio_fin and mes == mes_fin:
            break
        mes += 1
        if mes > 11:
            mes = 0
            anio += 1
    return periodos

def parse_periodo_input(periodo_input):
    periodo_input = periodo_input.strip()
    if not periodo_input:
        return []
    if '-' in periodo_input:
        parts = [p.strip() for p in periodo_input.split('-')]
        if len(parts) == 2:
            return generar_rango_periodos(parts[0], parts[1])
        return []
    if validar_periodo(periodo_input):
        return [periodo_input]
    return []

# ==================== LOGGING ====================
def log(msg, level="info"):
    ts = datetime.now().strftime("%H:%M:%S")
    entry = {"ts": ts, "msg": msg, "level": level}
    with job_lock:
        job_state["logs"].append(entry)
    if level == "error":
        logger.error(msg)
    else:
        logger.info(msg)

def reset_state():
    with job_lock:
        job_state["running"] = False
        job_state["stopping"] = False
        job_state["logs"] = []
        job_state["stats"] = {"total": 0, "descargadas": 0, "errores": 0}
        job_state["finished"] = False
        job_state["error"] = None
        job_state["errores_detalle"] = []
        job_state["descargas_exitosas"] = []
        job_state["facturas_permitidas"] = []

def stop_job():
    global current_browser, current_context, current_dl_dir, current_periodo, current_ips_nombre
    with job_lock:
        job_state["stopping"] = True
    log("🛑 Solicitando detención del proceso...", "warn")
    if current_browser:
        try:
            current_browser.close()
            log("  → Navegador cerrado por solicitud de stop.")
        except Exception as e:
            log(f"  → Error al cerrar navegador: {e}", "error")
    generar_zip_parcial()

def generar_zip_parcial():
    global current_dl_dir, current_periodo, current_ips_nombre
    if not current_dl_dir or not current_periodo or not current_ips_nombre:
        return
    ips_dir = current_dl_dir / current_ips_nombre
    if not ips_dir.exists():
        return
    with job_lock:
        exitosas = job_state["descargas_exitosas"].copy()
        errores = job_state["errores_detalle"].copy()
    excel_parcial_path = None
    if EXCEL_AVAILABLE:
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_name = f"reporte_parcial_{timestamp}.xlsx"
            excel_parcial_path = ips_dir / excel_name
            wb = openpyxl.Workbook()
            ws_exit = wb.active
            ws_exit.title = "Descargadas"
            ws_exit.append(["N° Factura", "Estado", "IPS", "Archivo Descargado", "Fecha/Hora"])
            for ex in exitosas:
                ws_exit.append([ex.get("factura"), ex.get("estado"), current_ips_nombre, ex.get("archivo"), ex.get("timestamp")])
            ws_err = wb.create_sheet("Errores")
            ws_err.append(["N° Factura", "Estado", "IPS", "Error", "Captura pantalla", "Fecha/Hora"])
            for err in errores:
                ws_err.append([err.get("factura"), err.get("estado"), current_ips_nombre, err.get("error"), err.get("captura"), err.get("timestamp")])
            wb.save(excel_parcial_path)
            log(f"📊 Reporte Excel parcial generado: {excel_parcial_path}")
        except Exception as e:
            log(f"⚠️ No se pudo generar Excel parcial: {e}", "warn")
    archivos_a_incluir = list(ips_dir.rglob("*.pdf"))
    if excel_parcial_path and excel_parcial_path.exists():
        archivos_a_incluir.append(excel_parcial_path)
    errores_dir = ips_dir / "Errores"
    if errores_dir.exists():
        archivos_a_incluir.extend(errores_dir.rglob("*"))
    if not archivos_a_incluir:
        return
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        zip_name = f"facturas_{current_periodo}_PARCIAL_{timestamp}.zip"
        zip_path = current_dl_dir / zip_name
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for archivo in archivos_a_incluir:
                arcname = archivo.relative_to(current_dl_dir)
                zf.write(archivo, arcname=str(arcname))
        log(f"📦 ZIP parcial generado: {zip_path}")
    except Exception as e:
        log(f"⚠️ No se pudo generar ZIP parcial: {e}", "warn")

def crear_zip_completo(dl_dir, periodo, ips_nombre):
    try:
        zip_final_name = f"facturas_{periodo}.zip"
        zip_final_path = dl_dir / zip_final_name
        with zipfile.ZipFile(zip_final_path, "w", zipfile.ZIP_DEFLATED) as zf:
            ips_dir = dl_dir / ips_nombre
            if ips_dir.exists():
                for pdf in ips_dir.rglob("*.pdf"):
                    zf.write(pdf, arcname=str(pdf.relative_to(dl_dir)))
                for excel in ips_dir.glob("reporte_*.xlsx"):
                    if "_PARCIAL_" not in excel.name:
                        zf.write(excel, arcname=str(excel.relative_to(dl_dir)))
                errores_dir = ips_dir / "Errores"
                if errores_dir.exists():
                    for err_file in errores_dir.rglob("*"):
                        zf.write(err_file, arcname=str(err_file.relative_to(dl_dir)))
        log(f"📦 ZIP final generado: {zip_final_path}")
        return str(zip_final_path)
    except Exception as e:
        log(f"⚠️ No se pudo generar el ZIP final: {e}", "warn")
        return None

def cargar_progreso(ips_dir):
    progreso_path = ips_dir / "progreso.json"
    if progreso_path.exists():
        try:
            with open(progreso_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                completadas = data.get("completadas", [])
                return set(completadas) if isinstance(completadas, list) else set()
        except Exception as e:
            log(f"⚠️ Error al leer progreso: {e}", "warn")
    return set()

def guardar_progreso(ips_dir, completadas):
    progreso_path = ips_dir / "progreso.json"
    try:
        data = {"completadas": list(completadas), "actualizado": datetime.now().isoformat()}
        with open(progreso_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
    except Exception as e:
        log(f"⚠️ Error al guardar progreso: {e}", "warn")

def generar_reporte_excel(dl_dir, periodo, ips_nombre, exitosas, errores):
    if not EXCEL_AVAILABLE:
        return None
    excel_path = dl_dir / ips_nombre / f"reporte_{periodo}.xlsx"
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws_exit = wb.active
    ws_exit.title = "Descargadas"
    ws_exit.append(["N° Factura", "Estado", "IPS", "Archivo Descargado", "Fecha/Hora"])
    for ex in exitosas:
        ws_exit.append([ex.get("factura"), ex.get("estado"), ips_nombre, ex.get("archivo"), ex.get("timestamp")])
    ws_err = wb.create_sheet("Errores")
    ws_err.append(["N° Factura", "Estado", "IPS", "Error", "Captura pantalla", "Fecha/Hora"])
    for err in errores:
        ws_err.append([err.get("factura"), err.get("estado"), ips_nombre, err.get("error"), err.get("captura"), err.get("timestamp")])
    wb.save(excel_path)
    return excel_path

# ==================== FUNCIONES AUXILIARES ====================
def _find_frame_with_text(page, regex_text):
    js = f"() => {{ const re = new RegExp({json.dumps(regex_text)}, 'i'); return re.test(document.body?.innerText || ''); }}"
    for fr in page.frames:
        try:
            if fr.evaluate(js):
                return fr
        except:
            continue
    return None

def _cerrar_traza_factura(page):
    js = """
        () => {
            const headers = document.querySelectorAll('.ui-dialog-titlebar, .modal-header, [class*="header"]');
            for (const h of headers) {
                if (h.textContent && h.textContent.includes('Traza de Factura')) {
                    const dlg = h.closest('.ui-dialog, .modal, [role="dialog"]');
                    if (dlg) {
                        const closeBtn = dlg.querySelector('.ui-dialog-titlebar-close, button.close, [aria-label*="lose"], [class*="close"]');
                        if (closeBtn) { closeBtn.click(); return true; }
                    }
                }
            }
            return false;
        }
    """
    for fr in page.frames:
        try:
            if fr.evaluate(js):
                time.sleep(0.5)
                return
        except:
            continue

def _extraer_nombre_ips(page, target_frame):
    def _buscar_en_frame(frame):
        try:
            nit = frame.evaluate("() => { const match = document.body.innerText.match(/NIT\\s*:\\s*([\\d\\-\\s]+)/i); if(match) return match[1].replace(/[^0-9]/g, ''); return ''; }").strip()
            return nit if nit else ""
        except:
            return ""
    nit = _buscar_en_frame(page) or _buscar_en_frame(target_frame)
    if not nit:
        for fr in page.frames:
            if fr not in (page, target_frame):
                nit = _buscar_en_frame(fr)
                if nit:
                    break
    log(f"    🔍 NIT detectado: {nit}")
    if nit and nit in MAPA_IPS:
        nombre = MAPA_IPS[nit]
        log(f"    🏥 Nombre obtenido del mapa: {nombre}")
        return nombre
    js_nombre = """
        () => {
            const keywords = ["IPS","CLINICA","HOSPITAL","CENTRO","FUNDACIÓN","URGENCIAS","SALUD","ODONTOTRANS","URGETRAUMA","CORDIALIDAD"];
            for (const el of document.querySelectorAll('h1,h2,h3,h4,p,div')) {
                let txt = el.innerText.trim();
                if (txt.length > 5 && txt.length < 100 && keywords.some(kw => txt.toUpperCase().includes(kw))) return txt;
            }
            return "";
        }
    """
    nombre = page.evaluate(js_nombre).strip() or target_frame.evaluate(js_nombre).strip()
    if not nombre:
        nombre = "IPS_DESCONOCIDA"
    nombre = re.sub(r'[\\/*?:"<>|]', "", nombre).strip().replace(" ", "_")
    log(f"    🏥 IPS final: {nombre} (NIT: {nit})")
    return nombre

# ==================== FUNCIÓN DE DESCARGA CON ESPERA DEL CONTADOR ====================
def _download_factura(page, context, modal_frame, fac: dict, dl_dir: Path, ips_nombre: str):
    import img2pdf
    from PIL import Image
    import io

    num = fac["num"]
    tipo = fac["tipo"]
    if tipo == "devolucion":
        target_label = "ActaDevolucion"
        target_label_norm = target_label.replace('ó', 'o').replace('í', 'i')
        subcarpeta = "Devolucion"
        nombre_soporte = "ActaDevolución"
    else:
        target_label = "Envios_D"
        target_label_norm = target_label.replace('í', 'i')
        subcarpeta = "Auditada"
        nombre_soporte = "Envios_D"

    ips_dir = dl_dir / ips_nombre
    dl_subdir = ips_dir / subcarpeta
    dl_subdir.mkdir(parents=True, exist_ok=True)

    bot_id = fac.get("botId")
    log(f"    🔗 Abriendo factura {num}...")
    num_solo_digitos = re.sub(r'\D', '', str(num))

    js_click_robusto = f"""
        () => {{
            const botId = '{bot_id}';
            const targetDigits = '{num_solo_digitos}';
            const fila = document.querySelector(`[data-bot-row-id="${{botId}}"]`);
            if (!fila) return {{ ok: false, reason: "fila_no_encontrada" }};
            fila.scrollIntoView({{block: 'center'}});
            function dispararClick(el) {{
                if (!el) return false;
                try {{ el.click(); }} catch (e) {{}}
                try {{ el.dispatchEvent(new MouseEvent('click', {{bubbles: true, cancelable: true, view: window}})); }} catch (e) {{}}
                return true;
            }}
            const candidatos = [];
            for (const a of fila.querySelectorAll('a')) {{
                const t = (a.textContent || '').trim();
                if (t.replace(/\\D/g, '') === targetDigits || candidatos.length === 0)
                    candidatos.push({{ tipo: 'a', el: a }});
            }}
            for (const el of fila.querySelectorAll('[onclick]')) {{
                if (!candidatos.find(c => c.el === el)) candidatos.push({{ tipo: 'onclick', el }});
            }}
            candidatos.push({{ tipo: 'fila', el: fila }});
            for (const td of fila.querySelectorAll('td')) candidatos.push({{ tipo: 'td', el: td }});
            for (const c of candidatos) dispararClick(c.el);
            return {{ ok: true, clickedWith: 'cascada', candidates: candidatos.length }};
        }}
    """
    result = None
    try:
        result = modal_frame.evaluate(js_click_robusto)
    except Exception as e:
        log(f"    ⚠️ Click falló: {e}", "warn")
    if not result or not result.get("ok"):
        for fr in page.frames:
            try:
                r = fr.evaluate(js_click_robusto)
                if r and r.get("ok"):
                    result = r
                    break
            except:
                continue
    if not result or not result.get("ok"):
        raise Exception(f"Click totalmente fallido para factura {num}.")
    log(f"    ✓ Click en factura {num} OK.")
    time.sleep(1.5)

    detalle_state = None
    detalle_frame = None
    for _ in range(60):
        if job_state.get("stopping"): return
        f = _find_frame_with_text(page, "Adjuntos por Factura")
        if f:
            try:
                has_traza = f.evaluate("() => /Traza de Factura/i.test(document.body?.innerText || '')")
                detalle_state = "traza" if has_traza else "adjuntos_directo"
            except:
                detalle_state = "adjuntos_directo"
            detalle_frame = f
            break
        f = _find_frame_with_text(page, "Traza de Factura")
        if f:
            detalle_state = "traza"
            detalle_frame = f
        time.sleep(0.5)
    if not detalle_frame:
        raise Exception("No apareció 'Traza de Factura' ni 'Adjuntos por Factura'.")
    time.sleep(1.5)
    log(f"    ✅ Detalle abierto (modo: {detalle_state}).")

    if detalle_state == "traza":
        log("    📑 Forzando cambio a pestaña 'Soportes'...")
        soportes_ok = False
        for intento in range(5):
            if job_state.get("stopping"): return
            for fr in page.frames:
                try:
                    has_tabs = fr.evaluate(r"""() => /Factura.*Detalles.*Soportes/i.test((document.body?.innerText || '').replace(/\n/g, ' '))""")
                    if has_tabs:
                        try:
                            fr.locator("text=Soportes").first.click(timeout=5000)
                            soportes_ok = True
                            break
                        except:
                            if fr.evaluate("""() => { for(const el of document.querySelectorAll('*')) if((el.textContent||'').trim() === 'Soportes'){ el.click(); return true; } return false; }"""):
                                soportes_ok = True
                                break
                except:
                    continue
            if soportes_ok:
                break
            time.sleep(1)
        if not soportes_ok:
            log("    ⚠️ No se pudo clickear Soportes", "warn")
        else:
            time.sleep(3)

    log("    ⏳ Esperando 'Adjuntos por Factura'...")
    adjuntos_frame = None
    for _ in range(90):
        if job_state.get("stopping"): return
        for fr in page.frames:
            try:
                if fr.evaluate("() => /Adjuntos por Factura|Buscar por.*Fecha/i.test(document.body?.innerText || '')"):
                    adjuntos_frame = fr
                    break
            except:
                continue
        if adjuntos_frame:
            break
        time.sleep(0.5)
    if not adjuntos_frame:
        raise Exception("No se encontró sección 'Adjuntos por Factura'.")
    for _ in range(35):
        if job_state.get("stopping"): return
        try:
            if not adjuntos_frame.evaluate("() => /Procesando Solicitud/i.test(document.body?.innerText || '')"):
                break
        except:
            pass
        time.sleep(1)
    time.sleep(1)
    log("    ✅ Adjuntos cargados.")

    search_frame = adjuntos_frame

    def _escribir_buscador(texto):
        search_frame.evaluate("""
            () => {
                const inputs = document.querySelectorAll('input');
                for (const input of inputs) {
                    if ((input.placeholder || '').toLowerCase().includes('buscar') || (input.placeholder || '').toLowerCase().includes('filtrar')) {
                        input.value = '';
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        break;
                    }
                }
            }
        """)
        time.sleep(0.5)
        search_frame.evaluate(f"""
            () => {{
                const target = '{texto.replace('í', 'i')}';
                let searchInput = null;
                for (const input of document.querySelectorAll('input')) {{
                    const ph = (input.placeholder || '').toLowerCase();
                    if (ph.includes('buscar') || ph.includes('filtrar')) {{
                        searchInput = input;
                        break;
                    }}
                }}
                if (!searchInput) return;
                searchInput.focus();
                searchInput.select();
                const nativeSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                if (nativeSetter) nativeSetter.call(searchInput, target);
                else searchInput.value = target;
                searchInput.dispatchEvent(new Event('input', {{ bubbles: true }}));
                searchInput.dispatchEvent(new Event('change', {{ bubbles: true }}));
                let parent = searchInput.closest('div, td, form, span');
                if (parent) {{
                    for (const btn of parent.querySelectorAll('button, a, [role="button"], span')) {{
                        const html = (btn.outerHTML || '').toLowerCase();
                        const title = (btn.title || '').toLowerCase();
                        if (html.includes('search') || html.includes('lup') || title.includes('search')) {{
                            btn.click();
                            return;
                        }}
                    }}
                }}
                for (const svg of document.querySelectorAll('svg')) {{
                    if ((svg.outerHTML || '').toLowerCase().includes('search')) {{
                        const container = svg.closest('button, a, [role="button"]');
                        if (container) {{ container.click(); return; }}
                    }}
                }}
                searchInput.dispatchEvent(new KeyboardEvent('keypress', {{ key: 'Enter', bubbles: true }}));
            }}
        """)
        time.sleep(2)
        for _ in range(40):
            if job_state.get("stopping"): return
            processing = any(fr.evaluate("() => /Procesando Solicitud/i.test(document.body?.innerText || '')") for fr in page.frames if fr)
            if not processing:
                break
            time.sleep(0.5)
        time.sleep(2)

    log(f"    🔍 Buscando '{target_label}'...")
    _escribir_buscador(target_label)
    archivo_seleccionado = False
    tipo_encontrado = None
    posibles_nombres = list({target_label, target_label_norm})

    for intento in range(4):
        if job_state.get("stopping"): return
        for fr in page.frames:
            try:
                resultado = fr.evaluate(f"""
                    () => {{
                        const nombres = {json.dumps(posibles_nombres)};
                        let contenedor = null;
                        for (const el of document.querySelectorAll('td, div, span, li, p, tr')) {{
                            const txt = (el.innerText || '').trim();
                            for (const nombre of nombres) {{
                                if (txt === nombre) {{
                                    contenedor = el.closest('div[class*="file"], li[class*="file"], tr, div[class*="item"], div[class*="attach"], div[class*="row"]');
                                    if (!contenedor) contenedor = el.closest('div, li, tr');
                                    break;
                                }}
                            }}
                            if (contenedor) break;
                        }}
                        if (!contenedor) return {{ ok: false }};
                        let check = contenedor.querySelector('input[type="checkbox"], input[type="radio"], [role="checkbox"]');
                        if (!check) check = contenedor.parentElement?.querySelector('input[type="checkbox"], input[type="radio"], [role="checkbox"]');
                        if (check) {{
                            if (!check.checked) {{
                                check.click();
                                check.checked = true;
                                check.dispatchEvent(new Event('change', {{ bubbles: true }}));
                            }}
                            return {{ ok: true, metodo: 'checkbox' }};
                        }}
                        let iconoPdf = contenedor.querySelector('img[src*="pdf"], svg[aria-label*="pdf"], i[class*="pdf"], i[class*="file"], div[class*="pdf-icon"]');
                        if (iconoPdf) {{
                            iconoPdf.click();
                            return {{ ok: true, metodo: 'icono_pdf' }};
                        }}
                        contenedor.click();
                        contenedor.dispatchEvent(new MouseEvent('click', {{ bubbles: true, cancelable: true }}));
                        contenedor.dispatchEvent(new MouseEvent('dblclick', {{ bubbles: true, cancelable: true }}));
                        return {{ ok: true, metodo: 'contenedor_forzado' }};
                    }}
                """)
                if resultado and resultado.get('ok'):
                    log(f"    ✅ Selección realizada (método: {resultado.get('metodo')})")
                    archivo_seleccionado = True
                    tipo_encontrado = nombre_soporte
                    break
            except Exception as e:
                log(f"    ⚠️ Error en intento {intento+1}: {e}", "warn")
        if archivo_seleccionado:
            break
        log(f"    🔄 Reintentando selección ({intento+1}/4)...")
        time.sleep(2)

    if not archivo_seleccionado:
        log(f"    ⚠️ No se encontró '{target_label}'. Intentando con 'Carta de'...")
        texto_busqueda = "Carta de"
        _escribir_buscador(texto_busqueda)
        archivo_seleccionado = False
        for intento in range(4):
            if job_state.get("stopping"): return
            for fr in page.frames:
                try:
                    resultado = fr.evaluate(f"""
                        () => {{
                            const buscarTexto = '{texto_busqueda}';
                            function normalizar(s) {{
                                return s.toLowerCase().normalize("NFD").replace(/[\\u0300-\\u036f]/g, "");
                            }}
                            for (const el of document.querySelectorAll('td, div, span, li, p, tr')) {{
                                const txt = (el.innerText || '').trim();
                                if (normalizar(txt).includes(normalizar(buscarTexto))) {{
                                    let contenedor = el.closest('div[class*="file"], li[class*="file"], tr, div[class*="item"], div[class*="attach"], div[class*="row"]');
                                    if (!contenedor) contenedor = el.closest('div, li, tr');
                                    if (contenedor) {{
                                        let check = contenedor.querySelector('input[type="checkbox"], input[type="radio"], [role="checkbox"]');
                                        if (!check) check = contenedor.parentElement?.querySelector('input[type="checkbox"], input[type="radio"], [role="checkbox"]');
                                        if (check) {{
                                            if (!check.checked) {{
                                                check.click();
                                                check.checked = true;
                                                check.dispatchEvent(new Event('change', {{ bubbles: true }}));
                                            }}
                                            return {{ ok: true, metodo: 'checkbox', texto: txt }};
                                        }}
                                        let iconoPdf = contenedor.querySelector('img[src*="pdf"], svg[aria-label*="pdf"], i[class*="pdf"]');
                                        if (iconoPdf) {{
                                            iconoPdf.click();
                                            return {{ ok: true, metodo: 'icono_pdf', texto: txt }};
                                        }}
                                        contenedor.click();
                                        contenedor.dispatchEvent(new MouseEvent('click', {{ bubbles: true, cancelable: true }}));
                                        contenedor.dispatchEvent(new MouseEvent('dblclick', {{ bubbles: true, cancelable: true }}));
                                        return {{ ok: true, metodo: 'contenedor_forzado', texto: txt }};
                                    }}
                                }}
                            }}
                            return {{ ok: false }};
                        }}
                    """)
                    if resultado and resultado.get('ok'):
                        log(f"    ✅ Selección realizada con '{texto_busqueda}' (método: {resultado.get('metodo')})")
                        archivo_seleccionado = True
                        tipo_encontrado = "CartaObjecion"
                        break
                except Exception as e:
                    log(f"    ⚠️ Error en intento {intento+1} para '{texto_busqueda}': {e}", "warn")
            if archivo_seleccionado:
                break
            log(f"    🔄 Reintentando '{texto_busqueda}' ({intento+1}/4)...")
            time.sleep(2)

    if not archivo_seleccionado:
        raise Exception(f"No se pudo seleccionar el archivo (intentó '{target_label}' y 'Carta de')")

    log("    ⏳ Esperando confirmación de selección...")
    for _ in range(20):
        if job_state.get("stopping"): return
        if not any(fr.evaluate("() => /Debe seleccionar por lo menos un documento/i.test(document.body?.innerText || '')") for fr in page.frames if fr):
            log("    ✅ Selección confirmada")
            break
        time.sleep(1)

    # ========== BOLÍVAR: ABRIR VISOR Y ESPERAR CONTADOR DE PÁGINAS ==========
    log(f"    👁️ Abriendo visor documental...")
    visor_page = None
    try:
        with context.expect_page(timeout=30000) as page_info:
            for fr in page.frames:
                try:
                    btn = fr.locator('button[title="Abrir Documento"], button[aria-label="Abrir Documento"], button:has(i.fa-eye), button:has(i.bi-eye)').first
                    if btn.is_visible(timeout=5000):
                        btn.click()
                        log("    ✅ Clic en botón 'Abrir Documento'")
                        break
                except:
                    pass
        visor_page = page_info.value
        time.sleep(2)
    except Exception as e:
        raise Exception(f"No se pudo abrir el visor documental: {e}")

    # === FUNCIÓN PARA ESPERAR EL CONTADOR DE PÁGINAS ===
    def _esperar_contador_paginas(frame, timeout=30):
        """Espera a que aparezca un texto con patrón 'X/Y' y retorna (actual, total) o None."""
        start = time.time()
        ultimo_valor = None
        while time.time() - start < timeout:
            if job_state.get("stopping"):
                return None
            # Buscar en todos los frames del visor
            for fr in [frame] + frame.frames:
                try:
                    # Buscar mediante regex en el texto del cuerpo
                    texto = fr.evaluate("() => document.body?.innerText || ''")
                    match = re.search(r'(\d+)\s*/\s*(\d+)', texto)
                    if match:
                        actual = int(match.group(1))
                        total = int(match.group(2))
                        # Si el total se mantiene estable durante 0.5s, lo damos por bueno
                        if ultimo_valor == (actual, total):
                            return (actual, total)
                        ultimo_valor = (actual, total)
                        time.sleep(0.5)
                        continue
                except:
                    pass
            time.sleep(0.5)
        return None

    log("    ⏳ Esperando contador de páginas (ej. '1/11')...")
    contador = _esperar_contador_paginas(visor_page, timeout=30)
    if not contador:
        log("    ⚠️ No se detectó contador de páginas. Se usará método alternativo (asumiendo 1 página).", "warn")
        total_paginas = 1
        pagina_actual_esperada = 1
    else:
        pagina_actual_esperada, total_paginas = contador
        log(f"    📄 Contador detectado: {pagina_actual_esperada}/{total_paginas}")

    # Función para capturar la imagen actual del documento (después de verificar contador)
    def _capturar_pagina_con_contador(frame, num_pagina_esperado, timeout=15):
        """Espera a que el contador muestre el número esperado, luego captura el contenido del documento."""
        start = time.time()
        while time.time() - start < timeout:
            if job_state.get("stopping"):
                return None
            # Verificar contador en todos los frames
            for fr in [frame] + frame.frames:
                try:
                    texto = fr.evaluate("() => document.body?.innerText || ''")
                    match = re.search(r'(\d+)\s*/\s*(\d+)', texto)
                    if match and int(match.group(1)) == num_pagina_esperado:
                        # El contador coincide, ahora capturar la imagen
                        # Intentar obtener canvas o imagen grande
                        for f in [fr] + fr.frames:
                            try:
                                canvas = f.locator("canvas").first
                                if canvas.is_visible():
                                    bbox = canvas.bounding_box()
                                    if bbox and bbox['width'] > 200:
                                        time.sleep(0.3)  # pequeña estabilización
                                        img_bytes = canvas.screenshot()
                                        if img_bytes and len(img_bytes) > 10000:
                                            return img_bytes
                            except:
                                pass
                            try:
                                imgs = f.locator("img")
                                for i in range(imgs.count()):
                                    img = imgs.nth(i)
                                    if img.is_visible():
                                        width = img.get_attribute("width") or 0
                                        if int(width) > 300:
                                            src = img.get_attribute("src")
                                            if src and (src.startswith("http") or src.startswith("blob")):
                                                resp = frame.request.get(src, timeout=10000)
                                                if resp.ok and len(resp.body()) > 10000:
                                                    return resp.body()
                                            img_bytes = img.screenshot()
                                            if img_bytes and len(img_bytes) > 10000:
                                                return img_bytes
                            except:
                                pass
                        # Fallback: screenshot de toda la página
                        screenshot = frame.screenshot()
                        if screenshot and len(screenshot) > 5000:
                            # Recortar bordes
                            try:
                                img = Image.open(io.BytesIO(screenshot))
                                w, h = img.size
                                left = int(w * 0.05)
                                top = int(h * 0.05)
                                right = int(w * 0.95)
                                bottom = int(h * 0.95)
                                img_cropped = img.crop((left, top, right, bottom))
                                output = io.BytesIO()
                                img_cropped.save(output, format='PNG')
                                return output.getvalue()
                            except:
                                return screenshot
                except:
                    pass
            time.sleep(0.5)
        return None

    imagenes_bytes = []
    if total_paginas == 1 and not contador:
        # Fallback: intentar capturar directamente
        log("    ⬇️ Intentando captura directa (sin contador)...")
        img_bytes = _capturar_pagina_con_contador(visor_page, 1, timeout=10)
        if img_bytes:
            imagenes_bytes.append(img_bytes)
            log("    ✅ Página capturada")
        else:
            log("    ⚠️ No se pudo capturar", "warn")
    else:
        log(f"    ⬇️ Iniciando descarga de {total_paginas} página(s) (esperando contador por página)...")
        for pagina in range(1, total_paginas + 1):
            if job_state.get("stopping"): return
            log(f"    📥 Esperando página {pagina}/{total_paginas}...")
            img_bytes = _capturar_pagina_con_contador(visor_page, pagina, timeout=20)
            if img_bytes:
                imagenes_bytes.append(img_bytes)
                log(f"    ✅ Página {pagina} capturada ({len(img_bytes)} bytes)")
            else:
                log(f"    ⚠️ No se pudo capturar página {pagina} (timeout)", "warn")
                # Intentar avanzar igualmente para no bloquear
                if pagina < total_paginas:
                    _avanzar_pagina(visor_page)
                    time.sleep(2)
            # Avanzar a la siguiente (si no es la última)
            if pagina < total_paginas:
                log(f"    ➡️ Avanzando a página {pagina+1}...")
                _avanzar_pagina(visor_page)
                time.sleep(1)  # dar tiempo a que el visor cambie

    try:
        visor_page.close()
    except:
        pass

    if not imagenes_bytes:
        raise Exception(f"No se capturó ninguna imagen para la factura {num}")

    log(f"    📦 Consolidando {len(imagenes_bytes)} imagen(es) en PDF...")
    soporte_encontrado = tipo_encontrado if tipo_encontrado else nombre_soporte
    safe_name = re.sub(r"[^\w\-_.]", "_", f"{num}_{soporte_encontrado}.pdf")
    out_path = dl_subdir / safe_name

    try:
        with open(out_path, "wb") as f:
            f.write(img2pdf.convert(imagenes_bytes))
        log(f"    💾 PDF guardado: {out_path.name} ({out_path.stat().st_size // 1024} KB)")
    except Exception as e:
        log(f"    ⚠️ img2pdf falló, usando PIL: {e}", "warn")
        try:
            pil_images = [Image.open(io.BytesIO(img)).convert('RGB') for img in imagenes_bytes]
            if pil_images:
                pil_images[0].save(str(out_path), save_all=True, append_images=pil_images[1:])
                log(f"    💾 PDF guardado (PIL): {out_path.name}")
        except Exception as e2:
            raise Exception(f"No se pudo consolidar PDF: {e2}")

    with job_lock:
        job_state["descargas_exitosas"].append({
            "factura": num,
            "estado": fac["estado"],
            "archivo": str(out_path),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
    _cerrar_traza_factura(page)
    time.sleep(0.8)

def _avanzar_pagina(page):
    """Avanza a la siguiente página en el visor de Bolívar."""
    for frame in page.frames:
        try:
            if frame.evaluate("""() => {
                const btns = document.querySelectorAll('button, a, [role="button"]');
                for (const btn of btns) {
                    const html = btn.outerHTML.toLowerCase();
                    const txt = (btn.textContent || '').toLowerCase();
                    if ((html.includes('arrow') || html.includes('chevron') || html.includes('right') ||
                         txt.includes('>') || txt.includes('next') || txt.includes('siguiente')) &&
                        !html.includes('left') && !html.includes('prev')) {
                        btn.click();
                        return true;
                    }
                }
                return false;
            }"""):
                time.sleep(0.5)
                return True
        except:
            pass
    try:
        page.evaluate("""() => {
            for (const el of document.querySelectorAll('button, div, span')) {
                const svgs = el.querySelectorAll('svg');
                for (const svg of svgs) {
                    const path = svg.outerHTML.toLowerCase();
                    if (path.includes('arrow') || path.includes('chevron')) {
                        el.click();
                        return;
                    }
                }
            }
        }""")
        time.sleep(0.5)
        return True
    except:
        return False

# ==================== AUTOMATIZACIÓN PRINCIPAL ====================
def run_automation(usuario: str, password: str, periodo: str, download_path: str):
    from playwright.sync_api import sync_playwright
    global current_browser, current_context, current_dl_dir, current_periodo, current_ips_nombre

    dl_dir = Path(download_path)
    dl_dir.mkdir(parents=True, exist_ok=True)
    ips_nombre_actual = "IPS_SIN_NOMBRE"
    zip_parcial_generado = False
    current_dl_dir = dl_dir
    current_periodo = periodo

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            context = browser.new_context(accept_downloads=True, viewport={"width": 1500, "height": 900})
            page = context.new_page()
            current_browser = browser
            current_context = context

            log("🔐 Iniciando sesión en Activa IT...")
            if job_state.get("stopping"): return
            page.goto("https://activa-it.net/Login.aspx", wait_until="networkidle", timeout=60000)
            page.fill('input[placeholder="Usuario"]', usuario)
            page.fill('input[placeholder="Contraseña"]', password)
            try:
                checkbox = page.locator('input[type="checkbox"]').first
                if not checkbox.is_checked():
                    checkbox.check()
            except:
                pass
            page.click('button:has-text("Inicio de sesión"), input[value="Inicio de sesión"]')
            page.wait_for_url("**/Index.aspx", timeout=60000)
            time.sleep(2)
            log("✅ Sesión iniciada correctamente.")
            if job_state.get("stopping"): return

            log("📂 Navegando a módulo BI IPS...")
            time.sleep(3)

            def _find_periodo_in_frames():
                js_check = f"""() => {{ const body = (document.body?.innerText || '').toLowerCase(); return body.includes('{periodo.lower()}') || ['abr26','abr-26','abr.26','abr/26','abr2026'].some(v => body.includes(v)); }}"""
                for fr in page.frames:
                    try:
                        if fr.evaluate(js_check):
                            return fr
                    except:
                        continue
                return None

            if job_state.get("stopping"): return
            clicked = False
            for intento in range(3):
                try:
                    page.locator("text=BI IPS").first.click(timeout=15000)
                    clicked = True
                    log("  ✓ Click directo en 'BI IPS' OK.")
                    break
                except:
                    pass
                try:
                    page.click("text=Inteligencia de Negocio", timeout=8000)
                    time.sleep(1)
                    page.click("text=BI IPS", timeout=8000)
                    clicked = True
                    log("  ✓ Click vía 'Inteligencia de Negocio' + 'BI IPS' OK.")
                    break
                except:
                    pass
                try:
                    page.click("[class*='menu-toggle'], [class*='hamburger'], .sidebar-toggle", timeout=5000)
                    time.sleep(2)
                    page.click("text=BI IPS", timeout=8000)
                    clicked = True
                    log("  ✓ Click vía hamburguesa + 'BI IPS' OK.")
                    break
                except Exception as e:
                    log(f"    ⚠️ Intento {intento+1} falló: {e}", "warn")
                    time.sleep(2)
            if not clicked:
                raise Exception("No se encontró el módulo BI IPS en el menú.")

            time.sleep(3)
            log("✅ Módulo BI IPS abierto. Buscando período...")
            target_frame = None
            for i in range(120):
                if job_state.get("stopping"): return
                target_frame = _find_periodo_in_frames()
                if target_frame:
                    log(f"✅ Período '{periodo}' detectado tras {(i+1)*0.5:.1f}s.")
                    break
                time.sleep(0.5)
            if not target_frame:
                raise Exception(f"No se pudo localizar el período '{periodo}' tras 60s.")

            log("🏥 Obteniendo nombre de la IPS...")
            ips_nombre_actual = _extraer_nombre_ips(page, target_frame)
            current_ips_nombre = ips_nombre_actual

            if job_state.get("stopping"): return
            log(f"📅 Click en columna Cant del período '{periodo}'...")
            click_result = target_frame.evaluate(f"""
                () => {{
                    for (const row of document.querySelectorAll('tr')) {{
                        const cells = row.querySelectorAll('td');
                        if (cells.length < 3) continue;
                        if (cells[0].textContent.trim() !== '{periodo}') continue;
                        const links = row.querySelectorAll('a');
                        if (links.length === 0) return {{ ok: false, reason: 'sin_links' }};
                        const firstLink = links[0];
                        const value = firstLink.textContent.trim();
                        if (value === '0') return {{ ok: false, reason: 'cant_cero', value: '0' }};
                        firstLink.scrollIntoView({{block: 'center'}});
                        firstLink.click();
                        return {{ ok: true, value: value }};
                    }}
                    return {{ ok: false, reason: 'fila_no_encontrada' }};
                }}
            """)
            if click_result.get("reason") == "cant_cero":
                log(f"ℹ️ El período '{periodo}' tiene 0 facturas radicadas.", "warn")
                browser.close()
                return
            if not click_result.get("ok"):
                raise Exception(f"No se pudo hacer click en Cant de '{periodo}': {click_result.get('reason')}")
            log(f"  → Click en Cant: {click_result.get('value')}")

            log("⏳ Esperando modal 'Listado de facturas recibidas'...")
            modal_frame = None
            for _ in range(60):
                if job_state.get("stopping"): return
                for fr in page.frames:
                    try:
                        if fr.evaluate("() => /Listado de facturas recibidas/i.test(document.body?.innerText || '')"):
                            modal_frame = fr
                            break
                    except:
                        continue
                if modal_frame:
                    break
                time.sleep(0.5)
            if not modal_frame:
                raise Exception("El modal 'Listado de facturas recibidas' no apareció.")

            log("⏳ Esperando datos del listado...")
            data_frame = None
            for _ in range(120):
                if job_state.get("stopping"): return
                for fr in page.frames:
                    try:
                        if fr.evaluate("() => /Pendiente de recibir Informaci|Devoluci[oó]n de entrada/i.test(document.body?.innerText || '')"):
                            data_frame = fr
                            break
                    except:
                        continue
                if data_frame:
                    break
                time.sleep(0.5)
            if not data_frame:
                log("⚠️ No se encontraron facturas con los estados objetivo.", "warn")
                browser.close()
                return

            log(f"✅ Datos detectados en frame '{data_frame.name or '(main)'}'.")
            time.sleep(2)

            # ========== EXTRACCIÓN MEJORADA DE FACTURAS ==========
            log("🔍 Extrayendo facturas (incluye números con #)...")
            js_extract = r"""
            (state) => {
                const ESTADOS = [
                    { nombre: 'Auditada: Pendiente de recibir Informacion', regex: /auditada\s*:\s*pendiente\s+de\s+recibir\s+informaci[oó]n/i, tipo: 'auditada' },
                    { nombre: 'En radicacion: Devolución de entrada', regex: /en\s+radicaci[oó]n\s*:\s*devoluci[oó]n\s+de\s+entrada/i, tipo: 'devolucion' },
                    { nombre: 'En auditoria: Pendiente de informar Orden de pago al Pagador', regex: /en\s+auditori?a\s*:\s*pendiente\s+de\s+informar\s+orden\s+de\s+pago\s+al\s+pagador/i, tipo: 'auditada' },
                ];
                function normalizar(s) {
                    return s.toLowerCase().normalize("NFD").replace(/[\u0300-\u036f]/g, "")
                        .replace(/[‘’´`]/g, "'").replace(/[“”]/g, '"');
                }
                const filas = document.querySelectorAll('tr, [role="row"], li');
                const nuevas = [];
                for (const fila of filas) {
                    let fullText = (fila.innerText || '').replace(/\s+/g, ' ').trim();
                    if (!fullText || fullText.length < 20 || fullText.length > 400) continue;
                    let tipoDetectado = null, nombreEstado = null;
                    const textoNorm = normalizar(fullText);
                    for (const e of ESTADOS) {
                        if (e.regex.test(textoNorm)) {
                            tipoDetectado = e.tipo;
                            nombreEstado = e.nombre;
                            break;
                        }
                    }
                    if (!tipoDetectado) continue;
                    const tokens = fullText.split(/\s+/);
                    let numFactura = null;
                    for (const token of tokens) {
                        let clean = token.replace(/^#+/, '');
                        let digits = clean.replace(/\D/g, '');
                        if (digits.length >= 6 && digits.length <= 12) {
                            numFactura = digits;
                            break;
                        }
                    }
                    if (!numFactura) continue;
                    if (state.seen.includes(numFactura)) continue;
                    const botId = 'bot_' + state.nextId;
                    state.nextId++;
                    fila.setAttribute('data-bot-row-id', botId);
                    nuevas.push({
                        botId: botId, num: numFactura, rawNum: numFactura,
                        tipo: tipoDetectado, estado: nombreEstado,
                        textoFila: fullText.slice(0, 150), tagName: fila.tagName.toLowerCase(),
                    });
                    state.seen.push(numFactura);
                }
                return { nuevas: nuevas, total: state.seen.length };
            }
            """
            extract_state = {"nextId": 0, "seen": []}
            facturas_acumuladas = []
            rondas_sin_nuevos = 0
            for ronda in range(20):
                if job_state.get("stopping"): return
                try:
                    res = data_frame.evaluate(js_extract, extract_state)
                except Exception as e:
                    log(f"  ⚠️ Error en extracción ronda {ronda+1}: {e}", "warn")
                    res = {"nuevas": []}
                nuevas = res.get("nuevas", [])
                if nuevas:
                    facturas_acumuladas.extend(nuevas)
                    rondas_sin_nuevos = 0
                    log(f"  Ronda {ronda+1}: +{len(nuevas)} (Total: {len(facturas_acumuladas)})")
                else:
                    rondas_sin_nuevos += 1
                extract_state["seen"] = list(set(extract_state["seen"] + [n["num"] for n in nuevas]))
                if rondas_sin_nuevos >= 5:
                    break
                try:
                    data_frame.evaluate("() => { const s = document.querySelectorAll('div, table, tbody, [class*=\"scroll\"]'); for (const e of s) if (e.scrollHeight > e.clientHeight + 20) e.scrollTop += e.clientHeight * 0.8; window.scrollBy(0, window.innerHeight * 0.8); }")
                except:
                    pass
                time.sleep(0.5)
            log(f"📊 {len(facturas_acumuladas)} facturas detectadas.")
            facturas_objetivo = facturas_acumuladas

            # ========== PERSISTENCIA Y FILTRO ==========
            ips_dir = dl_dir / ips_nombre_actual
            completadas = cargar_progreso(ips_dir)

            facturas_pendientes = []
            for fac in facturas_objetivo:
                if fac['num'] in completadas:
                    log(f"⏭️ Factura {fac['num']} ya descargada, omitiendo.")
                    with job_lock:
                        job_state["stats"]["descargadas"] += 1
                        job_state["descargas_exitosas"].append({
                            "factura": fac['num'],
                            "estado": fac['estado'],
                            "archivo": str(ips_dir / ("Auditada" if fac['tipo']=='auditada' else "Devolucion") / f"Factura_{fac['num']}_{('Envios_D' if fac['tipo']=='auditada' else 'ActaDevolucion')}.pdf"),
                            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        })
                else:
                    facturas_pendientes.append(fac)

            with job_lock:
                permitidas = job_state.get("facturas_permitidas", [])
            if permitidas:
                original_count = len(facturas_pendientes)
                facturas_pendientes = [fac for fac in facturas_pendientes if fac['num'] in permitidas]
                log(f"📋 Filtro activo: {len(facturas_pendientes)} de {original_count} facturas permitidas.")

            log(f"📋 Facturas pendientes: {len(facturas_pendientes)}")
            with job_lock:
                job_state["stats"]["total"] = len(facturas_pendientes) + job_state["stats"]["descargadas"]
                job_state["stats"]["errores"] = 0

            cnt_aud = sum(1 for f in facturas_pendientes if f["tipo"] == "auditada")
            cnt_dev = sum(1 for f in facturas_pendientes if f["tipo"] == "devolucion")
            log("📋 RESUMEN:")
            log(f"  • Auditada: {cnt_aud}")
            log(f"  • Devolucion: {cnt_dev}")
            log(f"  TOTAL: {len(facturas_pendientes)}")
            if not facturas_pendientes:
                log("ℹ️ No hay facturas pendientes.")
                browser.close()
                with job_lock:
                    exitosas = job_state["descargas_exitosas"].copy()
                    errores = job_state["errores_detalle"].copy()
                generar_reporte_excel(dl_dir, periodo, ips_nombre_actual, exitosas, errores)
                crear_zip_completo(dl_dir, periodo, ips_nombre_actual)
                return

            # ========== PROCESAR FACTURAS ==========
            for idx, fac in enumerate(facturas_pendientes, 1):
                if job_state.get("stopping"):
                    log("🛑 Proceso detenido por el usuario.")
                    if not zip_parcial_generado:
                        generar_zip_parcial()
                        zip_parcial_generado = True
                    return
                log(f"[{idx}/{len(facturas_pendientes)}] Factura {fac['num']} ({fac['tipo']})...")
                try:
                    _download_factura(page, context, data_frame, fac, dl_dir, ips_nombre_actual)
                    with job_lock:
                        job_state["stats"]["descargadas"] += 1
                    completadas.add(fac['num'])
                    guardar_progreso(ips_dir, completadas)
                    log(f"  ✅ Descargada: {fac['num']}", "success")
                except Exception as e:
                    with job_lock:
                        job_state["stats"]["errores"] += 1
                        error_msg = str(e)
                        if "seleccionar el archivo" in error_msg:
                            if fac['tipo'] == 'auditada':
                                error_msg = f"No se encontró soporte Envios_D ni Carta de Objecion"
                            else:
                                error_msg = f"No se encontró soporte ActaDevolucion ni Carta de Objecion"
                        error_info = {
                            "factura": fac['num'],
                            "estado": fac['estado'],
                            "error": error_msg,
                            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        }
                        try:
                            errores_dir = ips_dir / "Errores"
                            errores_dir.mkdir(parents=True, exist_ok=True)
                            cap_path = errores_dir / f"ERROR_{fac['num']}.png"
                            page.screenshot(path=str(cap_path))
                            error_info["captura"] = str(cap_path)
                        except:
                            error_info["captura"] = ""
                        job_state["errores_detalle"].append(error_info)
                    log(f"  ⚠️ Error: {error_msg}", "error")
                    _cerrar_traza_factura(page)
                    time.sleep(1)

            browser.close()
            with job_lock:
                exitosas = job_state["descargas_exitosas"].copy()
                errores = job_state["errores_detalle"].copy()
            generar_reporte_excel(dl_dir, periodo, ips_nombre_actual, exitosas, errores)
            crear_zip_completo(dl_dir, periodo, ips_nombre_actual)
            log("🎉 Proceso completado.")

    except Exception as e:
        if not job_state.get("stopping"):
            log(f"💥 Error crítico: {e}", "error")
            with job_lock:
                job_state["error"] = str(e)
        else:
            log("Proceso detenido por el usuario.")
        if not zip_parcial_generado:
            generar_zip_parcial()
    finally:
        with job_lock:
            job_state["running"] = False
            job_state["finished"] = True
            job_state["stopping"] = False
        current_browser = None
        current_context = None
        current_dl_dir = None
        current_periodo = None
        current_ips_nombre = None

# ==================== RUTAS FLASK (sin cambios, incluye /api/clean) ====================
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/start", methods=["POST"])
def start_job():
    data = request.json or {}
    usuario = data.get("usuario", "").strip()
    password = data.get("password", "").strip()
    periodo_input = data.get("periodo", "").strip()
    custom_path = data.get("download_path", "").strip()
    if not all([usuario, password, periodo_input]):
        return jsonify({"ok": False, "error": "Faltan campos requeridos"}), 400
    periodos = parse_periodo_input(periodo_input)
    if not periodos:
        return jsonify({"ok": False, "error": f"Formato de período inválido: '{periodo_input}'. Use MMMYY (ej: May26) o rango MMMYY-MMMYY"}), 400
    with job_lock:
        if job_state["running"]:
            return jsonify({"ok": False, "error": "Ya hay un proceso en ejecución"}), 409
        job_state["running"] = True
        job_state["finished"] = False
        job_state["error"] = None
        job_state["stats"] = {"total": 0, "descargadas": 0, "errores": 0}
        job_state["errores_detalle"] = []
        job_state["descargas_exitosas"] = []
    dl_path = custom_path if custom_path else str(DOWNLOAD_DIR / periodo_input)
    periodo_principal = periodos[0] if len(periodos) == 1 else periodo_input
    if len(periodos) > 1:
        log(f"📅 Procesando rango de {len(periodos)} períodos: {periodos[0]} → {periodos[-1]}")
        job_state["periodos_rango"] = periodos
    else:
        job_state["periodos_rango"] = None
    t = threading.Thread(target=run_automation, args=(usuario, password, periodo_principal, dl_path), daemon=True)
    t.start()
    return jsonify({"ok": True, "download_path": dl_path, "periodos_detectados": periodos})

@app.route("/api/stop", methods=["POST"])
def stop_job_route():
    with job_lock:
        if not job_state["running"]:
            return jsonify({"ok": False, "message": "No hay proceso en ejecución"}), 400
    stop_job()
    return jsonify({"ok": True, "message": "Deteniendo proceso..."})

@app.route("/api/reset", methods=["POST"])
def reset_job_route():
    data = request.json or {}
    periodo = data.get("periodo", "").strip()
    with job_lock:
        if job_state["running"]:
            stop_job()
            time.sleep(2)
    if periodo:
        periodo_dir = DOWNLOAD_DIR / periodo
        if periodo_dir.exists():
            for progreso_file in periodo_dir.glob("*/progreso.json"):
                try:
                    progreso_file.unlink()
                    log(f"🗑️ Progreso eliminado: {progreso_file}")
                except Exception as e:
                    log(f"⚠️ Error al borrar {progreso_file}: {e}", "warn")
        else:
            log(f"⚠️ No existe la carpeta del período '{periodo}'.", "warn")
    else:
        log("⚠️ No se especificó período, no se borró progreso.", "warn")
    reset_state()
    return jsonify({"ok": True, "message": "Estado reiniciado y progreso eliminado."})

@app.route("/api/clean", methods=["POST"])
def clean_downloads():
    data = request.json or {}
    periodo = data.get("periodo", "").strip()
    ips = data.get("ips", "").strip()
    if not periodo:
        return jsonify({"ok": False, "error": "Se requiere el parámetro 'periodo'"}), 400
    periodo_dir = DOWNLOAD_DIR / periodo
    if not periodo_dir.exists():
        return jsonify({"ok": False, "error": f"El período '{periodo}' no tiene datos descargados"}), 404
    def delete_directory(path):
        try:
            if path.exists():
                shutil.rmtree(path)
                return True
        except Exception as e:
            log(f"⚠️ Error al eliminar {path}: {e}", "error")
            return False
        return False
    eliminados = []
    if ips:
        ips_dir = periodo_dir / ips
        if ips_dir.exists():
            if delete_directory(ips_dir):
                eliminados.append(str(ips_dir))
            else:
                return jsonify({"ok": False, "error": f"No se pudo eliminar la carpeta de IPS '{ips}'"}), 500
        else:
            return jsonify({"ok": False, "error": f"No existe la IPS '{ips}' en el período '{periodo}'"}), 404
    else:
        if delete_directory(periodo_dir):
            eliminados.append(str(periodo_dir))
        else:
            return jsonify({"ok": False, "error": f"No se pudo eliminar la carpeta del período '{periodo}'"}), 500
    log(f"🧹 Limpieza completa realizada: {', '.join(eliminados)}")
    return jsonify({"ok": True, "message": f"Se eliminaron correctamente: {', '.join(eliminados)}", "eliminados": eliminados})

@app.route("/api/status")
def get_status():
    with job_lock:
        return jsonify({
            "running": job_state["running"],
            "finished": job_state["finished"],
            "error": job_state["error"],
            "stats": job_state["stats"],
            "logs": job_state["logs"][-200:],
        })

@app.route("/api/logs")
def get_logs():
    since = int(request.args.get("since", 0))
    with job_lock:
        return jsonify({"logs": job_state["logs"][since:]})

@app.route("/api/logs", methods=["DELETE"])
def clear_logs():
    with job_lock:
        job_state["logs"] = []
    return jsonify({"ok": True})

@app.route("/api/files")
def list_files():
    periodo = request.args.get("periodo", "")
    folder = DOWNLOAD_DIR / periodo if periodo else DOWNLOAD_DIR
    files = []
    if folder.exists():
        for f in sorted(folder.iterdir()):
            if f.is_file():
                files.append({"name": f.name, "size": f.stat().st_size, "path": str(f), "periodo": periodo})
    return jsonify({"files": files})

@app.route("/downloads/<path:filename>")
def download_file(filename):
    return send_from_directory(DOWNLOAD_DIR, filename, as_attachment=True)

@app.route("/api/periodos")
def get_periodos():
    periodos = []
    for d in DOWNLOAD_DIR.iterdir():
        if d.is_dir():
            count = len(list(d.glob("**/*.pdf")))
            periodos.append({"name": d.name, "count": count})
    return jsonify({"periodos": sorted(periodos, key=lambda x: x["name"], reverse=True)})

@app.route("/api/upload", methods=["POST"])
def upload_facturas():
    if 'file' not in request.files:
        return jsonify({"ok": False, "error": "No se envió ningún archivo"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"ok": False, "error": "Archivo vacío"}), 400
    try:
        filename = file.filename.lower()
        facturas = []
        if filename.endswith('.csv'):
            content = file.read().decode('utf-8')
            reader = csv.DictReader(content.splitlines())
            for row in reader:
                for col, val in row.items():
                    if 'factura' in col.lower():
                        facturas.append(val.strip())
                        break
        elif filename.endswith(('.xls', '.xlsx')):
            if not EXCEL_AVAILABLE:
                return jsonify({"ok": False, "error": "openpyxl no instalado"}), 500
            wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
            ws = wb.active
            col_idx = None
            for cell in ws[1]:
                if cell.value and 'factura' in str(cell.value).lower():
                    col_idx = cell.column
                    break
            if col_idx is None:
                return jsonify({"ok": False, "error": "No se encontró columna con 'factura'"}), 400
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[col_idx-1]
                if val:
                    facturas.append(str(val).strip())
        else:
            return jsonify({"ok": False, "error": "Formato no soportado. Use CSV o Excel"}), 400
        facturas_limpias = [re.sub(r'\D', '', f) for f in facturas if re.sub(r'\D', '', f)]
        if not facturas_limpias:
            return jsonify({"ok": False, "error": "No se encontraron números de factura válidos"}), 400
        with job_lock:
            job_state["facturas_permitidas"] = facturas_limpias
        log(f"📄 Se cargaron {len(facturas_limpias)} facturas desde el archivo.")
        return jsonify({"ok": True, "count": len(facturas_limpias), "facturas": facturas_limpias[:10]})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error al procesar archivo: {str(e)}"}), 500

@app.route("/api/progreso")
def get_progreso():
    periodo = request.args.get("periodo", "")
    ips = request.args.get("ips", "")
    if not periodo:
        return jsonify({"ok": False, "error": "Se requiere el parámetro 'periodo'"}), 400
    periodo_dir = DOWNLOAD_DIR / periodo
    if not periodo_dir.exists():
        return jsonify({"ok": True, "completadas": [], "mensaje": "No hay datos para este período"})
    if ips:
        ips_dir = periodo_dir / ips
        if not ips_dir.exists():
            return jsonify({"ok": False, "error": f"No existe la IPS '{ips}'"}), 404
    else:
        posibles = list(periodo_dir.iterdir())
        if not posibles:
            return jsonify({"ok": True, "completadas": [], "mensaje": "No hay subcarpetas de IPS"})
        ips_dir = None
        for d in posibles:
            if d.is_dir() and (d / "progreso.json").exists():
                ips_dir = d
                break
        if not ips_dir:
            ips_dir = posibles[0] if posibles[0].is_dir() else None
        if not ips_dir:
            return jsonify({"ok": True, "completadas": [], "mensaje": "No se encontró carpeta de IPS"})
        ips = ips_dir.name
    progreso_path = ips_dir / "progreso.json"
    if not progreso_path.exists():
        return jsonify({"ok": True, "completadas": [], "ips": ips, "mensaje": "Aún no hay facturas completadas"})
    try:
        with open(progreso_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        completadas = data.get("completadas", [])
        return jsonify({"ok": True, "completadas": completadas, "cantidad": len(completadas), "ips": ips, "actualizado": data.get("actualizado", "")})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error al leer progreso: {str(e)}"}), 500

@app.route("/api/exportar_progreso")
def exportar_progreso_excel():
    periodo = request.args.get("periodo", "")
    if not periodo:
        return jsonify({"ok": False, "error": "Se requiere el parámetro 'periodo'"}), 400
    if not EXCEL_AVAILABLE:
        return jsonify({"ok": False, "error": "openpyxl no instalado"}), 500
    periodo_dir = DOWNLOAD_DIR / periodo
    if not periodo_dir.exists():
        return jsonify({"ok": False, "error": f"No existe la carpeta del período '{periodo}'"}), 404
    progreso_files = list(periodo_dir.glob("*/progreso.json"))
    if not progreso_files:
        return jsonify({"ok": False, "error": f"No se encontró progreso.json para el período '{periodo}'"}), 404
    progreso_path = progreso_files[0]
    ips_nombre = progreso_path.parent.name
    try:
        with open(progreso_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        completadas = data.get("completadas", [])
        actualizado = data.get("actualizado", "")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas completadas"
        ws.append(["N° Factura", "Fecha de completado"])
        for factura in completadas:
            ws.append([factura, actualizado])
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"progreso_facturas_{periodo}_{ips_nombre}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error al generar Excel: {str(e)}"}), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    print("\n" + "=" * 55)
    print("  🏥 Activa IT — Descargador de Cartas Glosa")
    print("  🔷 Bolívar SOAT (Con espera explícita del contador de páginas)")
    print("=" * 55)
    print(f"  📂 Carpeta de descargas: {DOWNLOAD_DIR}")
    print(f"  🌐 Puerto: {port}")
    print("=" * 55 + "\n")
    app.run(host="0.0.0.0", port=port, debug=False)