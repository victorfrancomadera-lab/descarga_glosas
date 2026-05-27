"""
==============================================================
  SISTEMA DE GLOSAS - DESCARGA PORTAL ACTIVA IT
  Portal    : Activa IT (La Previsora SOAT)
  Desarrollado por: DESARROLLO E INNOVACION SALUD NET
==============================================================
"""

import os
import re
import json
import csv
import time
import threading
import logging
import zipfile
from datetime import datetime
from pathlib import Path
from io import BytesIO

try:
    import openpyxl
    from openpyxl.styles import Font
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_OK = True
except ImportError:
    PLAYWRIGHT_OK = False

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BASE_DIR     = Path(__file__).resolve().parent
DOWNLOAD_DIR = BASE_DIR / "output" / "portal_previsora"
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

# ==================== MAPA DE IPS POR NIT ====================
MAPA_IPS = {
    "900267064": "INVERSIONES_AZALUD_CLINICA_BAHIA",
    "900827065": "CENTRO_DE_DIAGNOSTICO_E_IMAGENES_BAHIA",
    "900657731": "CENTRO_MEDICO_Y_DE_REHABILITACION_BAHIA",
    "900826509": "RED_DE_URGENCIAS_DEL_MAGDALENA",
    "900513306": "FUNDACION_MARIA_REINA",
    "900600550": "INVERSIONES_MEDICAS_BARU",
    "900954800": "CENTRO_MEDICO_Y_DE_REHABILITACION_BARU",
    "900631361": "INVERSIONES_MEDICAS_VALLESALUD",
    "900257333": "ODONTOTRANS",
    "901081281": "URGETRAUMA",
    "900792417": "RED_DE_URGENCIAS_DE_LA_COSTA_PACIFICA",
    "901959993": "CLINICA_CORDIALIDAD",
    "900002780": "FUNDACION_CAMPBELL",
    "900900754": "CLINICA_VALLESALUD_SAN_FERNANDO",
    "901523868": "MOVID_IPS",
    "900558595": "FUNDACION_MEDICA_CAMPBELL",
    "802024329": "RED_DE_URGENCIAS_DE_LA_COSTA",
    "901149757": "UNIDAD_MEDICA_DE_TRAUMA_DEL_VALLE",
    "901057487": "TECNOLOGIA_DIAGNOSTICA_DEL_VALLE",
}

# ==================== ESTADO GLOBAL ====================
_job_state = {
    "running": False,
    "stopping": False,
    "logs": [],
    "stats": {"total": 0, "descargadas": 0, "errores": 0},
    "finished": False,
    "error": None,
    "errores_detalle": [],
    "descargas_exitosas": [],
    "facturas_permitidas": [],
    "zip_path": None,
    "excel_path": None,
}
_job_lock           = threading.Lock()
_current_browser    = None
_current_context    = None
_current_dl_dir     = None
_current_periodo    = None
_current_ips_nombre = None


# ==================== LOG Y ESTADO ====================
def _log(msg, level="info"):
    ts    = datetime.now().strftime("%H:%M:%S")
    entry = {"ts": ts, "msg": msg, "level": level}
    with _job_lock:
        _job_state["logs"].append(entry)
    if level == "error":
        logger.error(msg)
    else:
        logger.info(msg)


def get_state() -> dict:
    with _job_lock:
        return {
            "running":    _job_state["running"],
            "finished":   _job_state["finished"],
            "error":      _job_state["error"],
            "stats":      dict(_job_state["stats"]),
            "logs":       list(_job_state["logs"]),
            "zip_path":   _job_state["zip_path"],
            "excel_path": _job_state["excel_path"],
        }


def get_logs_since(since: int = 0) -> list:
    with _job_lock:
        return list(_job_state["logs"][since:])


def reset_state():
    with _job_lock:
        _job_state["running"]            = False
        _job_state["stopping"]           = False
        _job_state["logs"]               = []
        _job_state["stats"]              = {"total": 0, "descargadas": 0, "errores": 0}
        _job_state["finished"]           = False
        _job_state["error"]              = None
        _job_state["errores_detalle"]    = []
        _job_state["descargas_exitosas"] = []
        _job_state["facturas_permitidas"]= []
        _job_state["zip_path"]           = None
        _job_state["excel_path"]         = None


def set_facturas_permitidas(lista: list):
    with _job_lock:
        _job_state["facturas_permitidas"] = lista


# ==================== STOP ====================
def stop_portal():
    global _current_browser
    with _job_lock:
        _job_state["stopping"] = True
    _log("Solicitando detención del proceso...", "warn")
    if _current_browser:
        try:
            _current_browser.close()
            _log("  Navegador cerrado por solicitud de stop.")
        except Exception as e:
            _log(f"  Error al cerrar navegador: {e}", "error")
    _generar_zip_parcial()


# ==================== ZIP PARCIAL ====================
def _generar_zip_parcial():
    global _current_dl_dir, _current_periodo, _current_ips_nombre
    if not _current_dl_dir or not _current_periodo or not _current_ips_nombre:
        return
    ips_dir = _current_dl_dir / _current_ips_nombre
    if not ips_dir.exists():
        return

    with _job_lock:
        exitosas = _job_state["descargas_exitosas"].copy()
        errores  = _job_state["errores_detalle"].copy()

    excel_parcial_path = None
    if EXCEL_AVAILABLE:
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_parcial_path = ips_dir / f"reporte_parcial_{ts}.xlsx"
            wb   = openpyxl.Workbook()
            ws_e = wb.active
            ws_e.title = "Descargadas"
            ws_e.append(["N Factura", "Estado", "IPS", "Archivo Descargado", "Fecha/Hora"])
            for ex in exitosas:
                ws_e.append([ex.get("factura"), ex.get("estado"), _current_ips_nombre, ex.get("archivo"), ex.get("timestamp")])
            ws_err = wb.create_sheet("Errores")
            ws_err.append(["N Factura", "Estado", "IPS", "Error", "Captura pantalla", "Fecha/Hora"])
            for err in errores:
                ws_err.append([err.get("factura"), err.get("estado"), _current_ips_nombre, err.get("error"), err.get("captura"), err.get("timestamp")])
            wb.save(excel_parcial_path)
            _log(f"Reporte Excel parcial generado: {excel_parcial_path}")
        except Exception as e:
            _log(f"No se pudo generar Excel parcial: {e}", "warn")
            excel_parcial_path = None

    archivos = list(ips_dir.rglob("*.pdf"))
    if excel_parcial_path and excel_parcial_path.exists():
        archivos.append(excel_parcial_path)
    errores_dir = ips_dir / "Errores"
    if errores_dir.exists():
        archivos.extend(errores_dir.rglob("*"))

    if not archivos:
        return

    try:
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        zip_path = _current_dl_dir / f"facturas_{_current_periodo}_PARCIAL_{ts}.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for a in archivos:
                zf.write(a, arcname=str(a.relative_to(_current_dl_dir)))
        _log(f"ZIP parcial generado: {zip_path}")
        with _job_lock:
            _job_state["zip_path"] = str(zip_path)
    except Exception as e:
        _log(f"No se pudo generar ZIP parcial: {e}", "warn")


def _crear_zip_completo(dl_dir: Path, periodo: str, ips_nombre: str):
    try:
        zip_path = dl_dir / f"facturas_{periodo}.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            ips_dir = dl_dir / ips_nombre
            if ips_dir.exists():
                for pdf in ips_dir.rglob("*.pdf"):
                    zf.write(pdf, arcname=str(pdf.relative_to(dl_dir)))
                for excel in ips_dir.glob("reporte_*.xlsx"):
                    if "_PARCIAL_" not in excel.name:
                        zf.write(excel, arcname=str(excel.relative_to(dl_dir)))
                errores_dir = ips_dir / "Errores"
                if errores_dir.exists():
                    for ef in errores_dir.rglob("*"):
                        zf.write(ef, arcname=str(ef.relative_to(dl_dir)))
        _log(f"ZIP final generado: {zip_path}")
        return str(zip_path)
    except Exception as e:
        _log(f"No se pudo generar el ZIP final: {e}", "warn")
        return None


# ==================== PERSISTENCIA ====================
def _cargar_progreso(ips_dir: Path) -> set:
    p = ips_dir / "progreso.json"
    if p.exists():
        try:
            data        = json.loads(p.read_text(encoding="utf-8"))
            completadas = data.get("completadas", [])
            if isinstance(completadas, list):
                return set(completadas)
            elif isinstance(completadas, dict):
                return set(completadas.keys())
        except Exception as e:
            _log(f"Error al leer progreso: {e}", "warn")
    return set()


def _guardar_progreso(ips_dir: Path, completadas: set):
    p = ips_dir / "progreso.json"
    try:
        p.write_text(
            json.dumps({"completadas": list(completadas),
                        "actualizado": datetime.now().isoformat()}, indent=2),
            encoding="utf-8"
        )
    except Exception as e:
        _log(f"Error al guardar progreso: {e}", "warn")


def reset_progreso(periodo: str):
    periodo_dir = DOWNLOAD_DIR / periodo
    if not periodo_dir.exists():
        return
    for pf in periodo_dir.glob("*/progreso.json"):
        try:
            pf.unlink()
            _log(f"Progreso eliminado: {pf}")
        except Exception as e:
            _log(f"Error al borrar progreso: {e}", "warn")


# ==================== EXCEL ====================
def _generar_reporte_excel(dl_dir: Path, periodo: str, ips_nombre: str,
                            exitosas: list, errores: list):
    if not EXCEL_AVAILABLE:
        return None
    excel_path = dl_dir / ips_nombre / f"reporte_{periodo}.xlsx"
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb   = openpyxl.Workbook()
    ws_e = wb.active
    ws_e.title = "Descargadas"
    ws_e.append(["N Factura", "Estado", "IPS", "Archivo Descargado", "Fecha/Hora"])
    for ex in exitosas:
        ws_e.append([ex.get("factura"), ex.get("estado"), ips_nombre, ex.get("archivo"), ex.get("timestamp")])
    ws_err = wb.create_sheet("Errores")
    ws_err.append(["N Factura", "Estado", "IPS", "Error", "Captura pantalla", "Fecha/Hora"])
    for err in errores:
        ws_err.append([err.get("factura"), err.get("estado"), ips_nombre, err.get("error"), err.get("captura"), err.get("timestamp")])
    wb.save(excel_path)
    return excel_path


# ==================== AUXILIARES ====================
def _find_frame_with_text(page, regex_text: str):
    js = f"() => {{ const re = new RegExp({json.dumps(regex_text)}, 'i'); return re.test(document.body?.innerText || ''); }}"
    for fr in page.frames:
        try:
            if fr.evaluate(js):
                return fr
        except:
            continue
    return None


def _cerrar_traza_factura(page):
    js = """
        () => {
            const headers = document.querySelectorAll('.ui-dialog-titlebar, .modal-header, [class*="header"]');
            for (const h of headers) {
                if (h.textContent && h.textContent.includes('Traza de Factura')) {
                    const dlg = h.closest('.ui-dialog, .modal, [role="dialog"]');
                    if (dlg) {
                        const closeBtn = dlg.querySelector('.ui-dialog-titlebar-close, button.close, [aria-label*="lose"], [class*="close"]');
                        if (closeBtn) { closeBtn.click(); return true; }
                    }
                }
            }
            return false;
        }
    """
    for fr in page.frames:
        try:
            if fr.evaluate(js):
                time.sleep(0.5)
                return
        except:
            continue


def _extraer_nombre_ips(page, target_frame) -> str:
    def _buscar_nit(frame):
        try:
            js = """
                () => {
                    const body = document.body.innerText;
                    const match = body.match(/NIT\\s*:\\s*([\\d\\-\\s]+)/i);
                    if (match) { return match[1].replace(/[^0-9]/g, ''); }
                    return "";
                }
            """
            return frame.evaluate(js).strip()
        except:
            return ""

    nit = _buscar_nit(page) or _buscar_nit(target_frame)
    if not nit:
        for fr in page.frames:
            if fr not in (page, target_frame):
                nit = _buscar_nit(fr)
                if nit:
                    break

    _log(f"    NIT detectado: {nit}")

    if nit and nit in MAPA_IPS:
        nombre = MAPA_IPS[nit]
        _log(f"    IPS del mapa para NIT {nit}: {nombre}")
        return nombre

    js_nombre = """
        () => {
            const keywords = ["IPS","CLINICA","HOSPITAL","CENTRO","FUNDACION","URGENCIAS","SALUD","ODONTOTRANS","URGETRAUMA","CORDIALIDAD"];
            const elementos = document.querySelectorAll('h1,h2,h3,h4,p,div');
            for (const el of elementos) {
                let txt = el.innerText.trim();
                if (txt.length > 5 && txt.length < 100) {
                    for (let kw of keywords) {
                        if (txt.toUpperCase().includes(kw)) { return txt; }
                    }
                }
            }
            return "";
        }
    """
    nombre = ""
    try:
        nombre = page.evaluate(js_nombre).strip()
    except:
        pass
    if not nombre:
        try:
            nombre = target_frame.evaluate(js_nombre).strip()
        except:
            pass

    if not nombre:
        nombre = "IPS_DESCONOCIDA"

    nombre = re.sub(r'[\\/*?:"<>|]', "", nombre).strip().replace(" ", "_")
    _log(f"    IPS final: {nombre} (NIT: {nit})")
    return nombre


# ==================== DESCARGA DE UNA FACTURA ====================
def _download_factura(page, context, modal_frame, fac: dict, dl_dir: Path, ips_nombre: str):
    num               = fac["num"]
    tipo              = fac["tipo"]
    target_label      = "ActaDevolucion" if tipo == "devolucion" else "Envios_D"
    target_label_norm = target_label.replace('i', 'i')
    subcarpeta        = "Auditada" if tipo == "auditada" else "Devolucion"

    ips_dir        = dl_dir / ips_nombre
    dl_subdir      = ips_dir / subcarpeta
    dl_subdir.mkdir(parents=True, exist_ok=True)

    bot_id          = fac.get("botId")
    num_solo_digits = re.sub(r'\D', '', str(num))
    _log(f"    Abriendo factura {num}...")

    js_click = f"""
        () => {{
            const botId = '{bot_id}';
            const targetDigits = '{num_solo_digits}';
            const fila = document.querySelector(`[data-bot-row-id="${{botId}}"]`);
            if (!fila) return {{ ok: false, reason: "fila_no_encontrada" }};
            fila.scrollIntoView({{block: 'center'}});
            function dispararClick(el) {{
                if (!el) return false;
                try {{ el.click(); }} catch (e) {{}}
                try {{ el.dispatchEvent(new MouseEvent('click', {{bubbles: true, cancelable: true, view: window}})); }} catch (e) {{}}
                return true;
            }}
            const candidatos = [];
            for (const a of fila.querySelectorAll('a')) {{
                const t = (a.textContent || '').trim();
                if (t.replace(/\\D/g, '') === targetDigits || candidatos.length === 0)
                    candidatos.push({{ tipo: 'a', el: a }});
            }}
            for (const el of fila.querySelectorAll('[onclick]')) {{
                if (!candidatos.find(c => c.el === el)) candidatos.push({{ tipo: 'onclick', el }});
            }}
            candidatos.push({{ tipo: 'fila', el: fila }});
            for (const td of fila.querySelectorAll('td')) candidatos.push({{ tipo: 'td', el: td }});
            for (const c of candidatos) dispararClick(c.el);
            return {{ ok: true, clickedWith: 'cascada', candidates: candidatos.length }};
        }}
    """
    result = None
    try:
        result = modal_frame.evaluate(js_click)
    except Exception as e:
        _log(f"    Click fallo: {e}", "warn")
    if not result or not result.get("ok"):
        for fr in page.frames:
            try:
                r = fr.evaluate(js_click)
                if r and r.get("ok"):
                    result = r
                    break
            except:
                continue
    if not result or not result.get("ok"):
        raise Exception(f"Click totalmente fallido para factura {num}.")
    _log(f"    Click en factura {num} OK.")
    time.sleep(1.5)

    detalle_state = None
    detalle_frame = None
    for _ in range(60):
        if _job_state.get("stopping"): return
        f = _find_frame_with_text(page, "Adjuntos por Factura")
        if f:
            try:
                has_traza     = f.evaluate("() => /Traza de Factura/i.test(document.body?.innerText || '')")
                detalle_state = "traza" if has_traza else "adjuntos_directo"
            except:
                detalle_state = "adjuntos_directo"
            detalle_frame = f
            break
        f = _find_frame_with_text(page, "Traza de Factura")
        if f:
            detalle_state = "traza"
            detalle_frame = f
        time.sleep(0.5)
    if not detalle_frame:
        raise Exception("No aparecio 'Traza de Factura' ni 'Adjuntos por Factura'.")
    time.sleep(1.5)
    _log(f"    Detalle abierto (modo: {detalle_state}).")

    if detalle_state == "traza":
        _log("    Forzando cambio a pestana 'Soportes'...")
        soportes_ok = False
        for intento in range(5):
            if _job_state.get("stopping"): return
            for fr in page.frames:
                try:
                    has_tabs = fr.evaluate(r"""() => {
                        const txt = (document.body?.innerText || '').replace(/\n/g, ' ');
                        return /Factura.*Detalles.*Soportes/i.test(txt);
                    }""")
                    if has_tabs:
                        try:
                            fr.locator("text=Soportes").first.click(timeout=5000)
                            soportes_ok = True
                            break
                        except:
                            clicked = fr.evaluate("""() => {
                                for (const el of document.querySelectorAll('*')) {
                                    if ((el.textContent||'').trim() === 'Soportes') {
                                        el.click(); return true;
                                    }
                                }
                                return false;
                            }""")
                            if clicked:
                                soportes_ok = True
                                break
                except:
                    continue
            if soportes_ok:
                break
            time.sleep(1)
        if not soportes_ok:
            _log("    No se pudo clickear Soportes", "warn")
        else:
            time.sleep(3)

    _log("    Esperando 'Adjuntos por Factura'...")
    adjuntos_frame = None
    for _ in range(90):
        if _job_state.get("stopping"): return
        for fr in page.frames:
            try:
                if fr.evaluate("() => /Adjuntos por Factura|Buscar por.*Fecha/i.test(document.body?.innerText || '')"):
                    adjuntos_frame = fr
                    break
            except:
                continue
        if adjuntos_frame:
            break
        time.sleep(0.5)
    if not adjuntos_frame:
        raise Exception("No se encontro seccion 'Adjuntos por Factura'.")
    for _ in range(35):
        if _job_state.get("stopping"): return
        try:
            if not adjuntos_frame.evaluate("() => /Procesando Solicitud/i.test(document.body?.innerText || '')"):
                break
        except:
            pass
        time.sleep(1)
    time.sleep(1)
    _log("    Adjuntos cargados.")

    search_frame = adjuntos_frame

    def _escribir_buscador(texto):
        search_frame.evaluate("""
            () => {
                const inputs = document.querySelectorAll('input');
                for (const input of inputs) {
                    const ph = (input.placeholder || '').toLowerCase();
                    if (ph.includes('buscar') || ph.includes('filtrar') || ph.includes('nombre')) {
                        input.value = '';
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        break;
                    }
                }
            }
        """)
        time.sleep(0.5)
        search_frame.evaluate(f"""
            () => {{
                const target = '{texto.replace("i", "i")}';
                const inputs = document.querySelectorAll('input');
                let searchInput = null;
                for (const input of inputs) {{
                    const ph = (input.placeholder || '').toLowerCase();
                    if (ph.includes('buscar') || ph.includes('filtrar') || ph.includes('nombre')) {{
                        searchInput = input; break;
                    }}
                }}
                if (!searchInput) return;
                searchInput.focus();
                searchInput.select();
                const nativeSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                if (nativeSetter) nativeSetter.call(searchInput, target);
                else searchInput.value = target;
                searchInput.dispatchEvent(new Event('input', {{ bubbles: true }}));
                searchInput.dispatchEvent(new Event('change', {{ bubbles: true }}));
                let parent = searchInput.closest('div, td, form, span');
                if (parent) {{
                    const btns = parent.querySelectorAll('button, a, [role="button"], span');
                    for (const btn of btns) {{
                        const html = (btn.outerHTML || '').toLowerCase();
                        const title = (btn.title || '').toLowerCase();
                        if (html.includes('search') || html.includes('lup') || title.includes('search')) {{
                            btn.click(); return;
                        }}
                    }}
                }}
                const svgs = document.querySelectorAll('svg');
                for (const svg of svgs) {{
                    if ((svg.outerHTML || '').toLowerCase().includes('search')) {{
                        const container = svg.closest('button, a, [role="button"]');
                        if (container) {{ container.click(); return; }}
                    }}
                }}
                searchInput.dispatchEvent(new KeyboardEvent('keypress', {{ key: 'Enter', bubbles: true }}));
            }}
        """)
        time.sleep(2)
        for _ in range(40):
            if _job_state.get("stopping"): return
            processing = False
            for fr in page.frames:
                try:
                    if fr.evaluate("() => /Procesando Solicitud/i.test(document.body?.innerText || '')"):
                        processing = True
                        break
                except:
                    pass
            if not processing:
                break
            time.sleep(0.5)
        time.sleep(2)

    # ---------- PRIMERA BUSQUEDA: ENVIOS_D / ACTADEVOLUCION ----------
    _log(f"    Buscando '{target_label}'...")
    _escribir_buscador(target_label)
    archivo_seleccionado = False
    posibles_nombres     = list({target_label, target_label_norm})

    for intento in range(4):
        if _job_state.get("stopping"): return
        for fr in page.frames:
            try:
                resultado = fr.evaluate(f"""
                    () => {{
                        const nombres = {json.dumps(posibles_nombres)};
                        let contenedor = null;
                        const elementos = document.querySelectorAll('td, div, span, li, p, tr');
                        for (const el of elementos) {{
                            const txt = (el.innerText || '').trim();
                            for (const nombre of nombres) {{
                                if (txt === nombre) {{
                                    contenedor = el.closest('div[class*="file"], li[class*="file"], tr, div[class*="item"], div[class*="attach"], div[class*="row"]');
                                    if (!contenedor) contenedor = el.closest('div, li, tr');
                                    break;
                                }}
                            }}
                            if (contenedor) break;
                        }}
                        if (!contenedor) return {{ ok: false }};
                        let check = contenedor.querySelector('input[type="checkbox"], input[type="radio"], [role="checkbox"]');
                        if (!check) check = contenedor.parentElement?.querySelector('input[type="checkbox"], input[type="radio"], [role="checkbox"]');
                        if (check) {{
                            if (!check.checked) {{
                                check.click();
                                check.checked = true;
                                check.dispatchEvent(new Event('change', {{ bubbles: true }}));
                            }}
                            return {{ ok: true, metodo: 'checkbox' }};
                        }}
                        let iconoPdf = contenedor.querySelector('img[src*="pdf"], svg[aria-label*="pdf"], i[class*="pdf"], i[class*="file"], div[class*="pdf-icon"]');
                        if (iconoPdf) {{ iconoPdf.click(); return {{ ok: true, metodo: 'icono_pdf' }}; }}
                        contenedor.click();
                        contenedor.dispatchEvent(new MouseEvent('click', {{ bubbles: true, cancelable: true }}));
                        contenedor.dispatchEvent(new MouseEvent('dblclick', {{ bubbles: true, cancelable: true }}));
                        return {{ ok: true, metodo: 'contenedor_forzado' }};
                    }}
                """)
                if resultado and resultado.get('ok'):
                    _log(f"    Seleccion realizada (metodo: {resultado.get('metodo')})")
                    archivo_seleccionado = True
                    break
            except Exception as e:
                _log(f"    Error en intento {intento+1}: {e}", "warn")
        if archivo_seleccionado:
            break
        _log(f"    Reintentando seleccion ({intento+1}/4)...")
        time.sleep(2)

    # ---------- SEGUNDA BUSQUEDA: CARTA DE OBJECION ----------
    if not archivo_seleccionado:
        _log(f"    No se encontro '{target_label}'. Intentando con 'Carta de'...")
        texto_busqueda = "Carta de"
        _escribir_buscador(texto_busqueda)
        archivo_seleccionado = False
        for intento in range(4):
            if _job_state.get("stopping"): return
            for fr in page.frames:
                try:
                    resultado = fr.evaluate(f"""
                        () => {{
                            const buscarTexto = '{texto_busqueda}';
                            function normalizar(s) {{
                                return s.toLowerCase().normalize("NFD").replace(/[\\u0300-\\u036f]/g, "");
                            }}
                            const elementos = document.querySelectorAll('td, div, span, li, p, tr');
                            for (const el of elementos) {{
                                const txt = (el.innerText || '').trim();
                                if (normalizar(txt).includes(normalizar(buscarTexto))) {{
                                    let contenedor = el.closest('div[class*="file"], li[class*="file"], tr, div[class*="item"], div[class*="attach"], div[class*="row"]');
                                    if (!contenedor) contenedor = el.closest('div, li, tr');
                                    if (contenedor) {{
                                        let check = contenedor.querySelector('input[type="checkbox"], input[type="radio"], [role="checkbox"]');
                                        if (!check) check = contenedor.parentElement?.querySelector('input[type="checkbox"], input[type="radio"], [role="checkbox"]');
                                        if (check) {{
                                            if (!check.checked) {{
                                                check.click();
                                                check.checked = true;
                                                check.dispatchEvent(new Event('change', {{ bubbles: true }}));
                                            }}
                                            return {{ ok: true, metodo: 'checkbox', texto: txt }};
                                        }}
                                        let iconoPdf = contenedor.querySelector('img[src*="pdf"], svg[aria-label*="pdf"], i[class*="pdf"]');
                                        if (iconoPdf) {{ iconoPdf.click(); return {{ ok: true, metodo: 'icono_pdf', texto: txt }}; }}
                                        contenedor.click();
                                        contenedor.dispatchEvent(new MouseEvent('click', {{ bubbles: true, cancelable: true }}));
                                        contenedor.dispatchEvent(new MouseEvent('dblclick', {{ bubbles: true, cancelable: true }}));
                                        return {{ ok: true, metodo: 'contenedor_forzado', texto: txt }};
                                    }}
                                }}
                            }}
                            return {{ ok: false }};
                        }}
                    """)
                    if resultado and resultado.get('ok'):
                        _log(f"    Seleccion con 'Carta de' (metodo: {resultado.get('metodo')}) - Texto: '{resultado.get('texto')}'")
                        archivo_seleccionado = True
                        break
                except Exception as e:
                    _log(f"    Error en intento {intento+1}: {e}", "warn")
            if archivo_seleccionado:
                break
            _log(f"    Reintentando 'Carta de' ({intento+1}/4)...")
            time.sleep(2)

    if not archivo_seleccionado:
        raise Exception(f"No se pudo seleccionar el archivo (intento '{target_label}' y 'Carta de')")

    _log("    Esperando confirmacion de seleccion...")
    for _ in range(20):
        if _job_state.get("stopping"): return
        hay_error = False
        for fr in page.frames:
            try:
                if fr.evaluate("() => /Debe seleccionar por lo menos un documento/i.test(document.body?.innerText || '')"):
                    hay_error = True
                    break
            except:
                pass
        if not hay_error:
            _log("    Seleccion confirmada")
            break
        time.sleep(1)

    # ---------- ABRIR DOCUMENTO ----------
    _log(f"    Buscando boton 'Abrir Documento'...")
    pdf_data = None
    pdf_url  = None

    boton_encontrado = False
    start_time       = time.time()
    while time.time() - start_time < 15:
        if _job_state.get("stopping"): return
        for fr in page.frames:
            try:
                btn = fr.locator('button[title="Abrir Documento"], button[aria-label="Abrir Documento"], button:has(i.fa-eye), button:has(i.bi-eye)').first
                if btn.is_visible(timeout=2000):
                    boton_encontrado = True
                    break
            except:
                pass
        if boton_encontrado:
            break
        time.sleep(0.5)
    else:
        raise Exception("Boton 'Abrir Documento' no encontrado")

    for reintento in range(2):
        if _job_state.get("stopping"): return
        new_page = None
        try:
            with context.expect_page(timeout=30000) as page_info:
                for fr in page.frames:
                    try:
                        btn = fr.locator('button[title="Abrir Documento"], button[aria-label="Abrir Documento"], button:has(i.fa-eye), button:has(i.bi-eye)').first
                        if btn.is_visible(timeout=5000):
                            for _ in range(10):
                                if btn.is_enabled():
                                    break
                                time.sleep(0.5)
                            btn.click()
                            _log("    Clic en boton 'Abrir Documento'")
                            break
                    except:
                        pass
            new_page = page_info.value
            for _ in range(30):
                if _job_state.get("stopping"): return
                url = new_page.url
                if url and url != "about:blank" and ("amazonaws" in url or ".pdf" in url.lower()):
                    pdf_url = url
                    break
                time.sleep(0.5)
        except Exception as e:
            _log(f"    Intento {reintento+1}: No se abrio nueva pestana: {e}", "warn")
        finally:
            if new_page:
                try:
                    new_page.close()
                except:
                    pass

        if pdf_url:
            try:
                response = context.request.get(pdf_url, timeout=60000)
                if response.ok:
                    pdf_data = response.body()
                    _log(f"    PDF descargado ({len(pdf_data)//1024} KB)")
                    break
            except Exception as e:
                _log(f"    Error descargando: {e}", "warn")

        if not pdf_data:
            _log("    Intentando descarga directa...")
            try:
                with page.expect_download(timeout=30000) as download_info:
                    for fr in page.frames:
                        try:
                            btn = fr.locator('button[title="Abrir Documento"], button:has(i.fa-eye), button:has(i.bi-eye)').first
                            if btn.is_visible(timeout=3000):
                                btn.click()
                                break
                        except:
                            pass
                download = download_info.value
                pdf_data = download.path().read_bytes() if download.path() else None
                _log("    Descarga directa capturada")
                break
            except Exception as e:
                _log(f"    No se capturo descarga: {e}", "warn")

        if not pdf_data:
            _log(f"    Reintento {reintento+1}/2...")
            time.sleep(2)

    if not pdf_data:
        raise Exception("No se pudo obtener el PDF")

    safe_name = re.sub(r"[^\w\-_.]", "_", f"{num}_{target_label}.pdf")
    out_path  = dl_subdir / safe_name
    out_path.write_bytes(pdf_data)
    _log(f"    PDF guardado: {out_path.name} ({len(pdf_data)//1024} KB)")

    with _job_lock:
        _job_state["descargas_exitosas"].append({
            "factura":   num,
            "estado":    fac["estado"],
            "archivo":   str(out_path),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })

    _cerrar_traza_factura(page)
    time.sleep(0.8)


# ==================== AUTOMATIZACION PRINCIPAL ====================
def run_portal(usuario: str, password: str, periodo: str, dl_path=None):
    global _current_browser, _current_context, _current_dl_dir, _current_periodo, _current_ips_nombre

    if not PLAYWRIGHT_OK:
        _log("Playwright no esta instalado. Ejecute: playwright install chromium", "error")
        with _job_lock:
            _job_state["running"]  = False
            _job_state["finished"] = True
            _job_state["error"]    = "Playwright no disponible"
        return

    dl_dir = Path(dl_path) if dl_path else (DOWNLOAD_DIR / periodo)
    dl_dir.mkdir(parents=True, exist_ok=True)

    ips_nombre_actual    = "IPS_SIN_NOMBRE"
    zip_parcial_generado = False

    _current_dl_dir  = dl_dir
    _current_periodo = periodo

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            context = browser.new_context(accept_downloads=True, viewport={"width": 1500, "height": 900})
            page    = context.new_page()
            _current_browser = browser
            _current_context = context

            _log("Iniciando sesion en Activa IT...")
            if _job_state.get("stopping"): return
            page.goto("https://activa-it.net/Login.aspx", wait_until="networkidle", timeout=60000)
            _log(f"  Usuario: {usuario}")
            page.fill('input[placeholder="Usuario"]', usuario)
            page.fill('input[placeholder="Contraseña"]', password)
            try:
                checkbox = page.locator('input[type="checkbox"]').first
                if not checkbox.is_checked():
                    checkbox.check()
            except:
                pass
            page.click('button:has-text("Inicio de sesión"), input[value="Inicio de sesión"]')
            page.wait_for_url("**/Index.aspx", timeout=60000)
            time.sleep(2)
            _log("Sesion iniciada correctamente.")
            if _job_state.get("stopping"): return

            _log("Navegando a modulo BI IPS...")
            time.sleep(3)

            def _find_periodo_in_frames():
                js_check = f"""
                    () => {{
                        const bodyText = (document.body?.innerText || '').toLowerCase();
                        const periodo = '{periodo}'.toLowerCase();
                        if (bodyText.includes(periodo)) return true;
                        const variaciones = ['abr26', 'abr-26', 'abr.26', 'abr/26', 'abr2026'];
                        return variaciones.some(v => bodyText.includes(v));
                    }}
                """
                for fr in page.frames:
                    try:
                        if fr.evaluate(js_check):
                            return fr
                    except:
                        continue
                return None

            if _job_state.get("stopping"): return
            clicked = False
            for intento in range(3):
                try:
                    page.locator("text=BI IPS").first.click(timeout=15000)
                    clicked = True
                    _log("  Click directo en 'BI IPS' OK.")
                    break
                except:
                    pass
                try:
                    page.click("text=Inteligencia de Negocio", timeout=8000)
                    time.sleep(1)
                    page.click("text=BI IPS", timeout=8000)
                    clicked = True
                    _log("  Click via 'Inteligencia de Negocio' + 'BI IPS' OK.")
                    break
                except:
                    pass
                try:
                    page.click("[class*='menu-toggle'], [class*='hamburger'], .sidebar-toggle", timeout=5000)
                    time.sleep(2)
                    page.click("text=BI IPS", timeout=8000)
                    clicked = True
                    _log("  Click via hamburguesa + 'BI IPS' OK.")
                    break
                except Exception as e:
                    _log(f"    Intento {intento+1} fallo: {e}", "warn")
                    time.sleep(2)
            if not clicked:
                raise Exception("No se encontro el modulo BI IPS en el menu.")

            time.sleep(3)
            _log("Modulo BI IPS abierto. Buscando periodo...")
            target_frame = None
            for i in range(120):
                if _job_state.get("stopping"): return
                target_frame = _find_periodo_in_frames()
                if target_frame:
                    _log(f"Periodo '{periodo}' detectado tras {(i+1)*0.5:.1f}s.")
                    break
                time.sleep(0.5)
            if not target_frame:
                raise Exception(f"No se pudo localizar el periodo '{periodo}' tras 60s.")

            _log("Obteniendo nombre de la IPS...")
            ips_nombre_actual   = _extraer_nombre_ips(page, target_frame)
            _current_ips_nombre = ips_nombre_actual

            if _job_state.get("stopping"): return
            _log(f"Click en columna Cant del periodo '{periodo}'...")
            click_result = target_frame.evaluate(f"""
                () => {{
                    const rows = document.querySelectorAll('tr');
                    for (const row of rows) {{
                        const cells = row.querySelectorAll('td');
                        if (cells.length < 3) continue;
                        const firstText = cells[0].textContent.trim();
                        if (firstText !== '{periodo}') continue;
                        const links = row.querySelectorAll('a');
                        if (links.length === 0) return {{ ok: false, reason: 'sin_links' }};
                        const firstLink = links[0];
                        const value = firstLink.textContent.trim();
                        if (value === '0') return {{ ok: false, reason: 'cant_cero', value: '0' }};
                        firstLink.scrollIntoView({{block: 'center'}});
                        firstLink.click();
                        return {{ ok: true, value: value }};
                    }}
                    return {{ ok: false, reason: 'fila_no_encontrada' }};
                }}
            """)
            if click_result.get("reason") == "cant_cero":
                _log(f"El periodo '{periodo}' tiene 0 facturas radicadas.", "warn")
                browser.close()
                return
            if not click_result.get("ok"):
                raise Exception(f"No se pudo hacer click en Cant de '{periodo}': {click_result.get('reason')}")
            _log(f"  Click en Cant: {click_result.get('value')}")

            _log("Esperando modal 'Listado de facturas recibidas'...")
            modal_frame = None
            for _ in range(60):
                if _job_state.get("stopping"): return
                for fr in page.frames:
                    try:
                        if fr.evaluate("() => /Listado de facturas recibidas/i.test(document.body?.innerText || '')"):
                            modal_frame = fr
                            break
                    except:
                        continue
                if modal_frame:
                    break
                time.sleep(0.5)
            if not modal_frame:
                raise Exception("El modal 'Listado de facturas recibidas' no aparecio.")

            _log("Esperando datos del listado...")
            data_frame    = None
            tiempo_espera = 0
            while tiempo_espera < 60:
                if _job_state.get("stopping"): return
                for fr in page.frames:
                    try:
                        if fr.evaluate("() => /Pendiente de recibir Informaci|Devoluci[oo]n de entrada/i.test(document.body?.innerText || '')"):
                            data_frame = fr
                            break
                    except:
                        continue
                if data_frame:
                    break
                time.sleep(0.5)
                tiempo_espera += 0.5

            if not data_frame:
                _log("No se encontraron facturas con los estados objetivo.", "warn")
                browser.close()
                return

            _log(f"Datos detectados en frame '{data_frame.name or '(main)'}'.")
            time.sleep(2)

            _log("Extrayendo facturas...")
            js_extract = r"""
            (state) => {
                const ESTADOS = [
                    { nombre: 'Auditada: Pendiente de recibir Informacion', regex: /auditada\s*:\s*pendiente\s+de\s+recibir\s+informaci[oo]n/i, tipo: 'auditada' },
                    { nombre: 'En radicacion: Devolucion de entrada', regex: /en\s+radicaci[oo]n\s*:\s*devoluci[oo]n\s+de\s+entrada/i, tipo: 'devolucion' },
                    { nombre: 'En auditoria: Pendiente de informar Orden de pago al Pagador', regex: /en\s+auditori?a\s*:\s*pendiente\s+de\s+informar\s+orden\s+de\s+pago\s+al\s+pagador/i, tipo: 'auditada' },
                ];
                const filas = document.querySelectorAll('tr, [role="row"], li');
                const nuevas = [];
                for (const fila of filas) {
                    const fullText = (fila.innerText || '').replace(/\s+/g, ' ').trim();
                    if (!fullText || fullText.length < 20 || fullText.length > 400) continue;
                    if (!/\d{2}\/\d{2}\/\d{4}/.test(fullText)) continue;
                    let tipoDetectado = null, nombreEstado = null;
                    for (const e of ESTADOS) {
                        if (e.regex.test(fullText)) { tipoDetectado = e.tipo; nombreEstado = e.nombre; break; }
                    }
                    if (!tipoDetectado) continue;
                    const tokens = fullText.split(/\s+/);
                    const candidatosNum = tokens.filter(t => { const digits = t.replace(/\D/g, ''); return digits.length >= 6 && digits.length <= 10; });
                    if (candidatosNum.length === 0 || candidatosNum.length > 6) continue;
                    const numNorm = candidatosNum[0].replace(/\D/g, '');
                    if (state.seen.includes(numNorm)) continue;
                    const botId = 'bot_' + state.nextId;
                    state.nextId++;
                    fila.setAttribute('data-bot-row-id', botId);
                    nuevas.push({
                        botId: botId, num: numNorm, rawNum: candidatosNum[0],
                        tipo: tipoDetectado, estado: nombreEstado,
                        textoFila: fullText.slice(0, 150), tagName: fila.tagName.toLowerCase(),
                    });
                    state.seen.push(numNorm);
                }
                return { nuevas: nuevas, total: state.seen.length };
            }
            """
            extract_state       = {"nextId": 0, "seen": []}
            facturas_acumuladas = []
            rondas_sin_nuevos   = 0
            for ronda in range(20):
                if _job_state.get("stopping"): return
                try:
                    res = data_frame.evaluate(js_extract, extract_state)
                except:
                    res = {"nuevas": []}
                nuevas = res.get("nuevas", [])
                if nuevas:
                    facturas_acumuladas.extend(nuevas)
                    rondas_sin_nuevos = 0
                    _log(f"  Ronda {ronda+1}: +{len(nuevas)} (Total: {len(facturas_acumuladas)})")
                else:
                    rondas_sin_nuevos += 1
                extract_state["seen"] = list(set(extract_state["seen"] + [n["num"] for n in nuevas]))
                if rondas_sin_nuevos >= 5:
                    break
                try:
                    data_frame.evaluate("() => { const scrollables = document.querySelectorAll('div, table, tbody, [class*=\"scroll\"]'); for (const s of scrollables) { if (s.scrollHeight > s.clientHeight + 20) s.scrollTop += s.clientHeight * 0.8; } window.scrollBy(0, window.innerHeight * 0.8); }")
                except:
                    pass
                time.sleep(0.5)
            _log(f"{len(facturas_acumuladas)} facturas detectadas.")

            # ========== PERSISTENCIA Y FILTRO ==========
            ips_dir     = dl_dir / ips_nombre_actual
            completadas = _cargar_progreso(ips_dir)

            facturas_pendientes = []
            for fac in facturas_acumuladas:
                if fac['num'] in completadas:
                    _log(f"Factura {fac['num']} ya descargada, omitiendo.")
                    with _job_lock:
                        _job_state["stats"]["descargadas"] += 1
                        _job_state["descargas_exitosas"].append({
                            "factura":   fac['num'],
                            "estado":    fac['estado'],
                            "archivo":   str(ips_dir / ("Auditada" if fac['tipo']=='auditada' else "Devolucion") / f"{fac['num']}_{('Envios_D' if fac['tipo']=='auditada' else 'ActaDevolucion')}.pdf"),
                            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        })
                else:
                    facturas_pendientes.append(fac)

            with _job_lock:
                permitidas = _job_state.get("facturas_permitidas", [])
            if permitidas:
                original_count      = len(facturas_pendientes)
                facturas_pendientes = [fac for fac in facturas_pendientes if fac['num'] in permitidas]
                _log(f"Filtro activo: solo {len(facturas_pendientes)} de {original_count} facturas permitidas.")

            _log(f"Facturas pendientes: {len(facturas_pendientes)}")

            with _job_lock:
                _job_state["stats"]["total"]   = len(facturas_pendientes) + _job_state["stats"]["descargadas"]
                _job_state["stats"]["errores"] = 0

            cnt_aud = sum(1 for f in facturas_pendientes if f["tipo"] == "auditada")
            cnt_dev = sum(1 for f in facturas_pendientes if f["tipo"] == "devolucion")
            _log(f"  Auditada: {cnt_aud} | Devolucion: {cnt_dev} | TOTAL: {len(facturas_pendientes)}")

            if not facturas_pendientes:
                _log("No hay facturas pendientes por procesar.")
                browser.close()
                with _job_lock:
                    exitosas = _job_state["descargas_exitosas"].copy()
                    errores  = _job_state["errores_detalle"].copy()
                ep = _generar_reporte_excel(dl_dir, periodo, ips_nombre_actual, exitosas, errores)
                if ep:
                    with _job_lock:
                        _job_state["excel_path"] = str(ep)
                zp = _crear_zip_completo(dl_dir, periodo, ips_nombre_actual)
                if zp:
                    with _job_lock:
                        _job_state["zip_path"] = zp
                return

            # ========== PROCESAR FACTURAS ==========
            for idx, fac in enumerate(facturas_pendientes, 1):
                if _job_state.get("stopping"):
                    _log("Proceso detenido por el usuario.")
                    if not zip_parcial_generado:
                        _generar_zip_parcial()
                        zip_parcial_generado = True
                    return
                _log(f"[{idx}/{len(facturas_pendientes)}] Factura {fac['num']} ({fac['tipo']})...")
                try:
                    _download_factura(page, context, data_frame, fac, dl_dir, ips_nombre_actual)
                    with _job_lock:
                        _job_state["stats"]["descargadas"] += 1
                    completadas.add(fac['num'])
                    _guardar_progreso(ips_dir, completadas)
                    _log(f"  Descargada: {fac['num']}", "success")
                except Exception as e:
                    with _job_lock:
                        _job_state["stats"]["errores"] += 1
                        error_msg = str(e)
                        if "No se pudo seleccionar el archivo" in error_msg:
                            error_msg = f"En la factura {fac['num']} no se encontro soporte {('Envios_D' if fac['tipo']=='auditada' else 'ActaDevolucion')}"
                        error_info = {
                            "factura":   fac['num'],
                            "estado":    fac['estado'],
                            "error":     error_msg,
                            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "captura":   "",
                        }
                        try:
                            errores_dir = ips_dir / "Errores"
                            errores_dir.mkdir(parents=True, exist_ok=True)
                            cap_path    = errores_dir / f"ERROR_{fac['num']}.png"
                            page.screenshot(path=str(cap_path))
                            error_info["captura"] = str(cap_path)
                        except:
                            pass
                        _job_state["errores_detalle"].append(error_info)
                    _log(f"  Error: {error_msg}", "error")
                    _cerrar_traza_factura(page)
                    time.sleep(1)

            browser.close()

            with _job_lock:
                exitosas = _job_state["descargas_exitosas"].copy()
                errores  = _job_state["errores_detalle"].copy()
            ep = _generar_reporte_excel(dl_dir, periodo, ips_nombre_actual, exitosas, errores)
            if ep:
                _log(f"Reporte Excel generado: {ep}")
                with _job_lock:
                    _job_state["excel_path"] = str(ep)
            else:
                _log("No se pudo generar el Excel.", "warn")

            zp = _crear_zip_completo(dl_dir, periodo, ips_nombre_actual)
            if zp:
                with _job_lock:
                    _job_state["zip_path"] = zp

            _log("Proceso completado.", "success")

    except Exception as e:
        if not _job_state.get("stopping"):
            _log(f"Error critico: {e}", "error")
            with _job_lock:
                _job_state["error"] = str(e)
        else:
            _log("Proceso detenido por el usuario.")
        if not zip_parcial_generado:
            _generar_zip_parcial()
    finally:
        with _job_lock:
            _job_state["running"]  = False
            _job_state["finished"] = True
            _job_state["stopping"] = False
        _current_browser    = None
        _current_context    = None
        _current_dl_dir     = None
        _current_periodo    = None
        _current_ips_nombre = None
